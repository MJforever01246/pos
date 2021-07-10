import tkinter as tk
from tkinter import*
from tkinter import ttk
from tkinter import messagebox as mb
from openpyxl import load_workbook
from FullScreen import FullScreenApp
from OTP import GenerateOTP
from Mail_User_Profile import SendUserProfile
import hashlib
class EmployeesManagement(tk.Tk):
    def __init__(self):
        global tree
        # create self
        tk.Tk.__init__(self)
        FullScreenApp(self)
        self.title("Quản lí nhân viên")
        canvas=Canvas(master=self,width=1920,height=1080)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="employeemanagement.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        # table
        TableMargin = Frame(self, width=500)
        TableMargin.place(x=350, y=90)
        scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
        scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
        tree = ttk.Treeview(TableMargin,
                            columns=("Username", "Name", "DOB", "Gender", "Email", "Phone number", "Position"),
                            height=42, selectmode="extended", yscrollcommand=scrollbary.set,
                            xscrollcommand=scrollbarx.set)
        scrollbary.config(command=tree.yview)
        scrollbary.pack(side=RIGHT, fill=Y)
        scrollbarx.config(command=tree.xview)
        scrollbarx.pack(side=BOTTOM, fill=X)
        tree.heading('Username', text="Tên đăng nhập")
        tree.heading('Name', text="Tên")
        tree.heading('DOB', text="DOB (mm/dd/yyyy)")
        tree.heading('Gender', text="Giới tính")
        tree.heading('Email', text="Email")
        tree.heading('Phone number', text="Di động")
        tree.heading('Position', text="Chức vụ")
        tree.column('#0', stretch=NO, minwidth=0, width=0)
        tree.column('#1', stretch=NO, minwidth=0, width=240)
        tree.column('#2', stretch=NO, minwidth=0, width=240)
        tree.column('#3', stretch=NO, minwidth=0, width=240)
        tree.column('#4', stretch=NO, minwidth=0, width=80)
        tree.column('#5', stretch=NO, minwidth=0, width=240)
        tree.column('#6', stretch=NO, minwidth=0, width=240)
        tree.column('#7', stretch=NO, minwidth=0, width=240)
        tree.pack()
        excelfile = load_workbook("UsersData.xlsx")
        for sheet in excelfile.worksheets:
            for row in sheet.rows:
                if row[0].value == "Username" and row[1].value == "Password":
                    pass
                else:
                    tree.insert("", 0, values=(
                    row[0].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value))
        excelfile.close()
        # Search
        Search = Button(self, command=lambda: self.OpenSearchWindow("search"))
        Search.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        search_img = PhotoImage(file="SearchButton.png", master=canvas)
        Search.configure(image=search_img)
        Search.place(x=62, y=400)
        AddEmployee = Button(self,command=self.OpenRegisterWindow)
        AddEmployee.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                              foreground="#ffffff", background="#ffffff", borderwidth="0")
        AddEmployee_img = PhotoImage(file="AddEmployee.png", master=canvas)
        AddEmployee.configure(image=AddEmployee_img)
        AddEmployee.place(x=62, y=500)
        Edit = Button(self, command=lambda: self.OpenSearchWindow("Edit"))
        Edit.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                       foreground="#ffffff", background="#ffffff", borderwidth="0")
        Edit_img = PhotoImage(file="EditButton.png", master=canvas)
        Edit.configure(image=Edit_img)
        Edit.place(x=62, y=600)
        Delete = Button(self, command=lambda: self.OpenSearchWindow("Delete"))
        Delete.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        Delete_img = PhotoImage(file="DeleteAccount.png", master=canvas)
        Delete.configure(image=Delete_img)
        Delete.place(x=62, y=700)
        self.mainloop()
    def OpenSearchWindow(self,reason):
        global window_search
        window_search=SearchWindow(reason)
    def OpenRegisterWindow(self):
        window_register = RegisterWindowForUser()
class SearchWindow(tk.Tk):
    def __init__(self,reason):
        global Email_Phone, window_search
        tk.Tk.__init__(self)
        self.geometry("400x200+400+400")
        self.title("Tìm kiếm nhân viên")
        self.resizable(0, 0)
        back_search = Frame(master=self, bg='white')
        back_search.pack_propagate(0)
        back_search.pack(expand=1)
        canvas1 = Canvas(master=self, width=400, height=200, bg="#000000")
        bg_search = Label(master=self)
        bg_search.place(relx=0, rely=0, width=400, height=200)
        bg_img = PhotoImage(master=canvas1, width=400, height=200, file="SearchCustomer_Employee.png")
        bg_search.configure(image=bg_img)
        Email_Phone = Entry(master=self, relief="flat")
        Email_Phone.place(x=80, y=90, height=24, width=290)
        Search = Button(self, command=lambda: self.SearchWindowProcess(reason))
        Search.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        Search_img = PhotoImage(file="Search.png", master=canvas1)
        Search.configure(image=Search_img)
        Search.place(x=135, y=150)
        self.mainloop()
    def SearchWindowProcess(self,reason): #There are three cases which use this function: Edit, Register, Delete Account
        global list_profile,i
        Email_Phone_info = Email_Phone.get()
        self.destroy()
        excelfile=load_workbook("UsersData.xlsx")
        status = ""
        OK = False
        i = -2
        if "@" in Email_Phone_info or "." in Email_Phone_info:
            status = "Email"
        for sheet in excelfile.worksheets:
            for row in sheet.rows:
                i+=1
                if status == "Email":
                    if row[5].value == Email_Phone_info:
                        OK = True
                        if reason == "Edit":
                            list_profile=[]
                            for w in range(0,8):
                                list_profile.append(row[w].value)
                        break
                else:
                    if row[6].value == Email_Phone_info:
                        OK = True
                        if reason == "Edit":
                            list_profile=[]
                            for w in range(0,8):
                                list_profile.append(row[w].value)
                        break
        if OK == False:
            mb.showerror("Thông báo","Không tìm thấy thông tin")
        else:
            if reason == "Delete":
                self.remove_acc()
                mb.showinfo("Thông báo","Xóa tài khoản thành công")
            else:
                mb.showinfo("Thông báo", "Tìm kiếm thông tin thành công")
            if reason == "Edit":
                EditProfile()
            else:
                tree.selection_set(tree.get_children("")[-i-1])
                OK=False
        excelfile.close()
    def remove_acc(self):
        excelfile=load_workbook("UsersData.xlsx")
        data=excelfile["UsersData"]
        data.delete_rows(idx=i+2,amount=1)
        excelfile.save("UsersData.xlsx")
        excelfile.close()
        selected_item=tree.get_children("")[-i-1]
        tree.delete(selected_item)
class EditProfile(tk.Tk):
    def __init__(self):
        global name_entry, dob, option_gender,option_position, email_entry, phone_entry, username_entry, password_entry, position_combobox,status_hide,unhide_img,hide_img,hide,RadioButtonAdmin,RadioButtonCashier, order
        tk.Tk.__init__(self)
        self.geometry("1150x600+400+400")
        self.title("Chỉnh sửa tài khoản")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas = Canvas(master=self, width=1150, height=600)
        bg = Label(self)
        bg_img = PhotoImage(master=canvas, file="edituser.png")
        bg.configure(image=bg_img)
        bg.place(relx=0, rely=0)
        # name
        print(list_profile)
        name_entry = Entry(self, width=55, relief="flat")
        name_entry.place(x=85, y=126)
        name_entry.insert(END, list_profile[2])
        # DOB
        dob = Entry(self, width=55, relief="flat")
        dob.place(x=85, y=215)
        dob.insert(END,list_profile[3])
        # Gender
        option_gender = StringVar(master=self)
        option_gender.set(list_profile[4])
        img_male = PhotoImage(master=canvas, file="RadioButtonMale.png")
        RadioButtonMale = Checkbutton(master=self, image=img_male, bg="#ffffff", variable=option_gender, cursor="hand2",
                                      activebackground="#ffffff", onvalue="Nam",
                                      command=lambda: self.Checkboxgender("Nam"),offvalue="")
        RadioButtonMale.place(x=140, y=299)
        img_female = PhotoImage(master=canvas, file="RadioButtonFemale.png")
        RadioButtonFemale = Checkbutton(master=self, image=img_female, bg="#ffffff", variable=option_gender,
                                        cursor="hand2", activebackground="#ffffff", onvalue="Nữ",
                                        command=lambda: self.Checkboxgender("Nữ"),offvalue="")
        RadioButtonFemale.place(x=276, y=299)
        # email
        email_entry = Entry(self, width=55, relief="flat")
        email_entry.place(x=703, y=126)
        email_entry.insert(END, list_profile[5])
        # phone
        phone_entry = Entry(self, width=55, relief="flat")
        phone_entry.place(x=703, y=215)
        phone_entry.insert(END, list_profile[6])
        # username
        username_entry = Entry(self, width=55, relief="flat")
        username_entry.place(x=703, y=305)
        username_entry.insert(END, list_profile[0])
        # position
        option_position = tk.StringVar(master=self)
        option_position.set(list_profile[7])
        admin_img = PhotoImage(master=canvas, file="RadioButtonAdmin.png")
        RadioButtonAdmin = Checkbutton(master=self, image=admin_img, bg="#ffffff", variable=option_position,
                                       cursor="hand2", activebackground="#ffffff",
                                       command=lambda: self.Checkboxposition("Admin"), onvalue="Admin")
        RadioButtonAdmin.place(x=140, y=389)
        cashier_img = PhotoImage(master=canvas, file="RadioButtonCashier.png")
        RadioButtonCashier = Checkbutton(master=self, image=cashier_img, bg="#ffffff", variable=option_position,
                                         cursor="hand2", activebackground="#ffffff",
                                         command=lambda: self.Checkboxposition("Thu ngân"), onvalue="Thu ngân")
        RadioButtonCashier.place(x=276, y=389)
        # button
        complete_img = PhotoImage(master=canvas, file="Save.png")
        Complete = Button(self, command=self.process, image=complete_img, relief="flat", overrelief="flat",
                          activebackground="#ffffff", cursor="hand2",
                          foreground="#ffffff", background="#ffffff", borderwidth="0")
        Complete.place(x=505, y=550)
        order = i
        self.mainloop()
    def Checkboxposition(self, option):
        option_position.set(option)
    def Checkboxgender(self, option):
        option_gender.set(option)
    def process(self):
        global email_info, phone_info, status, otp
        status = True
        email_info=email_entry.get()
        phone_info=phone_entry.get()
        status_name=True
        status_email = True
        status_phone=True
        status_username=True
        status_position=True
        status_dob=True
        excelfile = load_workbook("UsersData.xlsx")
        user_data = excelfile["UsersData"]
        if list_profile[0] == username_entry.get():
            pass
        else:
            for cell in user_data["A"]:
                if cell.value == list_profile[0]:
                    pass
                elif cell.value == username_entry.get():
                    status = False
                    status_username = False
                    break
        if list_profile[5] == email_info:
            pass
        else:
            for cell in user_data["F"]:
                if cell.value == list_profile[5]:
                    pass
                elif cell.value == email_info:
                    status = False
                    status_email = False
                    break
        if name_entry.get()=="":
            status=False
            status_name=False
        if list_profile[6] == phone_info:
            pass
        else:
            for cell in user_data["G"]:
                if cell.value == list_profile[6]:
                    pass
                elif cell.value == phone_info:
                    status = False
                    status_phone = False
                    break
        excelfile.close()
        if status_email == True:
            if email_info.find("@") == -1 or email_info.find(".") == -1:
                status=False
                status_email=False
        if status_phone == True:
            for i in phone_info:
                if i in ["0","1","2","3","4","5","6","7","8","9"]:
                    pass
                else:
                    status_phone = False
                    status = False
                    break
            if status_phone == True:
                if len(phone_info) == 10 and phone_info[0] == "0":
                    pass
                else:
                    status=False
                    status_phone=False
        if status_username == True:
            if len(username_entry.get()) <= 3:
                status = False
                status_username=False
        if option_position.get() == "Admin" or option_position.get() == "Thu ngân":
            pass
        else:
            status_position=False
            status=False
        if dob.get() == "":
            pass
        else:
            cache = dob.get()
            cache = cache.split("/")
            try:
                if int(cache[0]) > 0 and int(cache[0]) < 32 and int(cache[1]) > 0 and int(cache[1]) < 13 and int(
                        cache[2]) > 1900 and int(cache[2]) < 2021:
                    pass
                else:
                    status = False
                    status_dob = False
            except:
                status = False
                status_dob = False
        if status==False:
            info=""
            if status_name == False:
                if info == "":
                    info += "Tên"
                else:
                    info += ", Tên"
            if status_dob == False:
                if info == "":
                    info += "Ngày sinh"
                else:
                    info += ", Ngày sinh"
            if status_position==False:
                if info=="":
                    info += "Chức vụ"
                else:
                    info+=", Chức vụ"
            if status_phone==False:
                if info=="":
                    info += "Di động"
                else:
                    info+=", Di động"
            if status_email==False:
                if info=="":
                    info += "Email"
                else:
                    info+=", Email"
            if status_username==False:
                if info=="":
                    info += "Tên đăng nhập"
                else:
                    info+=", Tên đăng nhập"
            mb.showerror("Thông báo",info+" không hợp lệ")
        if status == True:
            excelfile = load_workbook("UsersData.xlsx")
            for sheet in excelfile.worksheets:
                for row in sheet.rows:
                    if row[0].value == list_profile[0]:
                        row[0].value =str(username_entry.get())
                        row[2].value =str(name_entry.get())
                        row[3].value=str(dob.get())
                        row[4].value=str(option_gender.get())
                        row[5].value=str(email_info)
                        row[6].value=str(phone_info)
                        row[7].value=str(option_position.get())
                        excelfile.save("UsersData.xlsx")
                        excelfile.close()
                        break
            selected_item=tree.get_children("")[-order-1]
            tree.delete(selected_item)
            tree.insert("", -order-1, values=(str(username_entry.get()),str(name_entry.get()),str(dob.get()),str(option_gender.get()),str(email_info),str(phone_info),str(option_position.get())))
            self.destroy()
            mb.showinfo("Thông báo","Thay đổi thông tin thành công")
class RegisterWindowForUser(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        global name_entry, dob, option_gender,option_position, email_entry, phone_entry, username_entry, password_entry, position_combobox,status_hide,unhide_img,hide_img,hide,RadioButtonAdmin,RadioButtonCashier
        # create window
        self.geometry("1150x600+400+400")
        self.title("Đăng kí tài khoản")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="registerwindow.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        # name
        name_entry = Entry(self, width=55,relief="flat")
        name_entry.place(x=85, y=126)
        # DOB
        dob = Entry(self,width=55,relief="flat")
        dob.place(x=85, y=215)
        # Gender
        option_gender=StringVar()
        img_male=PhotoImage(master=canvas,file="RadioButtonMale.png")
        RadioButtonMale=Checkbutton(master=self,image=img_male,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nam",command=lambda:self.Checkboxgender("Nam"))
        RadioButtonMale.place(x=140, y=299)
        img_female=PhotoImage(master=canvas,file="RadioButtonFemale.png")
        RadioButtonFemale=Checkbutton(master=self,image=img_female,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nữ",command=lambda:self.Checkboxgender("Nữ"))
        RadioButtonFemale.place(x=276, y=299)
        # email
        email_entry = Entry(self, width=55,relief="flat")
        email_entry.place(x=703, y=126)
        # phone
        phone_entry = Entry(self, width=55,relief="flat")
        phone_entry.place(x=703, y=215)
        # username
        username_entry = Entry(self, width=55,relief="flat")
        username_entry.place(x=703, y=305)
        # password
        password_entry = Entry(self, width=55, show="-",relief="flat")
        password_entry.place(x=703, y=393)
        hide_img=PhotoImage(master=canvas,file="HidePassword.png")
        unhide_img=PhotoImage(master=canvas,file="UnhidePassword.png")
        hide=Button(self,command=self.hide_unhide_password,image=hide_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        status_hide=True
        hide.place(x=1059,y=389)
        # position
        option_position=tk.StringVar()
        admin_img=PhotoImage(master=canvas,file="RadioButtonAdmin.png")
        RadioButtonAdmin=Checkbutton(master=self,image=admin_img,bg="#ffffff",variable=option_position,cursor="hand2",activebackground="#ffffff",command=lambda:self.Checkboxposition("Admin"),onvalue="Admin")
        RadioButtonAdmin.place(x=140,y=389)
        cashier_img = PhotoImage(master=canvas, file="RadioButtonCashier.png")
        RadioButtonCashier = Checkbutton(master=self, image=cashier_img, bg="#ffffff", variable=option_position,cursor="hand2",activebackground="#ffffff",command=lambda:self.Checkboxposition("Thu ngân"),onvalue="Thu ngân")
        RadioButtonCashier.place(x=276, y=389)
        # button
        complete_img=PhotoImage(master=canvas,file="Save.png")
        Complete = Button(self,command=self.process,image=complete_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        Complete.place(x=505, y=550)
        self.mainloop()
    def hide_unhide_password(self):
        global status_hide
        if status_hide==True:
            hide.configure(image=unhide_img)
            password_entry.configure(show="")
            status_hide=False
        else:
            hide.configure(image=hide_img)
            password_entry.configure(show="-")
            status_hide = True
    def Checkboxposition(self,option):
        option_position.set(option)
    def Checkboxgender(self,option):
        option_gender.set(option)
    def process(self):
        global email_info, phone_info, status, otp, window_otp,gender,position
        status = True
        email_info = email_entry.get()
        phone_info = phone_entry.get()
        status_name=True
        status_email = True
        status_phone = True
        status_username = True
        status_position = True
        status_password = True
        status_dob=True
        excelfile = load_workbook("UsersData.xlsx")
        user_data = excelfile.worksheets[0]
        for row in user_data.rows:
            if row[0].value == username_entry.get():
                status = False
                status_username = False
            if row[5].value == email_info:
                status = False
                status_email = False
            if row[6].value == phone_info:
                status = False
                status_phone = False
        excelfile.close()
        if status_email == True:
            if email_info.find("@") == -1 or email_info.find(".") == -1:
                status = False
                status_email = False
        if status_phone == True:
            for i in phone_info:
                if i in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                    pass
                else:
                    status_phone = False
                    status = False
                    break
            if status_phone == True:
                if len(phone_info) == 10 and phone_info[0] == "0":
                    pass
                else:
                    status = False
                    status_phone = False
        if name_entry.get()=="":
            status_name=False
            status=False
        if dob.get() == "":
            pass
        else:
            cache = dob.get()
            cache = cache.split("/")
            try:
                if int(cache[0]) > 0 and int(cache[0]) < 32 and int(cache[1]) > 0 and int(cache[1]) < 13 and int(
                        cache[2]) > 1900 and int(cache[2]) < 2021:
                    pass
                else:
                    status = False
                    status_dob = False
            except:
                status = False
                status_dob = False
        if status_username == True:
            if len(username_entry.get()) <= 3:
                status = False
                status_username = False
        if option_position.get()=="Admin" or option_position.get()=="Thu ngân":
            pass
        else:
            status_position=False
            status=False
        if password_entry.get() == "":
            status = False
            status_password = False
        if status == False:
            info = ""
            if status_name == False:
                if info == "":
                    info += "Tên"
                else:
                    info += ", Tên"
            if status_dob == False:
                if info == "":
                    info += "Ngày sinh"
                else:
                    info += ", Ngày sinh"
            if status_position == False:
                if info == "":
                    info += "Chức vụ"
                else:
                    info += ", Chức vụ"
            if status_phone == False:
                if info == "":
                    info += "Di động"
                else:
                    info += ", Di động"
            if status_email == False:
                if info == "":
                    info += "Email"
                else:
                    info += ", Email"
            if status_username == False:
                if info == "":
                    info += "Tên đăng nhập"
                else:
                    info += ", Tên đăng nhập"
            if status_password == False:
                if info == "":
                    info += "Mật khẩu"
                else:
                    info += ", Mật khẩu"
            mb.showerror("Thông báo", info + " không hợp lệ",master=self)
        if status == True:
            try:
                otp = GenerateOTP(email_info)
                window_otp = OTP_window(otp,str(username_entry.get()), str(password_entry.get()),str(name_entry.get()),str(dob.get()),str(option_gender.get()),email_info,phone_info, str(option_position.get()),self)
            except:
                mb.showerror("Thông báo", "Kiểm tra lại email đã nhập hoặc mạng internet", master=self)
class OTP_window(tk.Tk):
    def __init__(self,otp,username,password,name,dob,gender,email,phone,position,window):
        global OTP_entry, window_otp
        tk.Tk.__init__(self)
        # window
        self.geometry("400x200+400+400")
        self.title("OTP")
        self.resizable(0, 0)
        canvas=Canvas(master=self,width=400,height=200)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="otp.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        OTP_entry = Entry(master=self,relief="flat")
        OTP_entry.place(x=80, y=92,height=24,width=292)
        Check_button = Button(self, command=lambda:self.checkOTP(otp,username,password,name,dob,gender,email,phone,position,window))
        checkbutton_img=PhotoImage(master=canvas,file="OTPCheck.png")
        Check_button.configure(image=checkbutton_img)
        Check_button.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                       foreground="#ffffff", background="#ffffff", borderwidth="0")
        Check_button.place(x=135, y=150)
        self.mainloop()
    def checkOTP(self,otp,username,password,name,dob,gender,email,phone,position,window):
        if OTP_entry.get() == otp:
            try:
                SendUserProfile(name,dob,gender,email,phone,username,position)
            except:
                pass
            self.SaveUserInfo(username,password,name,dob,gender,email,phone,position,window)
            self.destroy()
        else:
            ask=mb.askyesno("Thông báo", "Mã OTP của bạn không trùng khớp. Bạn có muốn thử lại hay không?",master=self)
            if ask==True:
                pass
            if ask==False:
                window.destroy()
                self.destroy()
                mb.showerror("Thông báo","Đăng kí tài khoản không thành công")
    def SaveUserInfo(self,username,password,name,dob,gender,email,phone,position,window):
        excelfile = load_workbook("UsersData.xlsx")
        data=excelfile["UsersData"]
        savedata = [username,str(hashlib.sha512(password.encode()).hexdigest()),name,dob,gender,email,phone,position]
        data.insert_rows(idx=2, amount=1)
        list_order = ["A2", "B2", "C2", "D2", "E2", "F2", "G2", "H2"]
        for i in range(8):
            data[list_order[i]] = savedata[i]
        excelfile.save("UsersData.xlsx")
        mb.showinfo("Thông báo","Thêm nhân viên thành công",master=window)
        window.destroy()
        savedata.pop(1)
        tree.insert("",0,values=tuple(savedata))
        excelfile.close()
# EmployeesManagement()