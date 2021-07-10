from tkinter import *
from tkinter import messagebox
import tkinter as tk
from openpyxl import load_workbook
import random
class NewProduct(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        global name_entry, barcode_entry
        # create window
        self.geometry("1150x600+400+400")
        self.title("Thêm sản phẩm mới")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="addproduct.png",width=1150,height=600)
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0,width=1150,height=600)
        # name
        name_entry = Entry(self, width=56,relief="flat")
        name_entry.place(x=82.5, y=125)
        # Category
        category_entry = Entry(self, width=56,relief="flat")
        category_entry.place(x=701, y=391)
        # Unit
        unit_entry = Entry(self, width=56,relief="flat")
        unit_entry.place(x=82.5, y=213)
        # Quantity
        quantity_entry = Entry(self,width=56,relief="flat")
        quantity_entry.place(x=82.5,y=302)
        # Price
        price_entry = Entry(self, width=56,relief="flat")
        price_entry.place(x=701, y=213)
        # Cost
        cost_entry = Entry(self, width=56,relief="flat")
        cost_entry.place(x=701, y=302)
        # Barcode
        barcode_entry = Entry(self, width=47,relief="flat")
        barcode_entry.place(x=753, y=124)
        barcode_auto_create=Button(self,command=self.CreateBarcode,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        create_img=PhotoImage(master=canvas,file="Create.png")
        barcode_auto_create.configure(image=create_img)
        barcode_auto_create.place(x=698,y=121)
        # Brand
        brand_entry = Entry(self, width=56,relief="flat")
        brand_entry.place(x=82.5, y=391)
        # button
        Complete = Button(self,command=lambda:self.CheckValid(name_entry.get(),quantity_entry.get(),unit_entry.get(),category_entry.get(),price_entry.get(),cost_entry.get(),barcode_entry.get(),brand_entry.get()),relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        buttom_img=PhotoImage(master=canvas,file="Save.png")
        Complete.configure(image=buttom_img)
        Complete.place(x=500, y=528)
        self.mainloop()
    def SaveProductInfo(self,name,quantity,unit,category,price,cost,barcode,brand):
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        list_order=["A2","B2","C2","D2","E2","F2","G2","H2"]
        data_list=[name,category,int(quantity),unit,int(price),int(cost),barcode,brand]
        data.insert_rows(idx=2,amount=1)
        for i in range (8):
            data[list_order[i]].value=data_list[i]
        data["F2"].number_format = '#,##0'
        data["E2"].number_format = '#,##0'
        excelfile.save("ProductsList.xlsx")
        excelfile.close()
        messagebox.showinfo("Thông báo","Thêm sản phầm thành công")
        self.destroy()
    def CheckValid(self,name,quantity,unit,category,price,cost,barcode,brand):
        status=True
        status_name=True
        status_unit=True
        status_category=True
        status_quantity=True
        status_price=True
        status_cost=True
        status_barcode=True
        if name == "":
            status_name=False
            status=False
        if quantity.isdigit():
            pass
        else:
            status = False
            status_quantity=False
        if unit=="":
            status=False
            status_unit=False
        if category == "":
            status=False
            status_category=False
        if price.isdigit():
            pass
        else:
            status=False
            status_price=False
        if cost.isdigit():
            pass
        else:
            status=False
            status_cost=False
        if barcode == "":
            status = False
            status_barcode=False
        if status_barcode==True or status_name==True:
            excelfile=load_workbook("ProductsList.xlsx")
            data=excelfile["ProductsList"]
            if status_barcode==True:
                for cell in data["G"]:
                    if cell == barcode:
                        status_barcode=False
                        status=False
                        break
            if status_name==True:
                for cell in data["A"]:
                    if cell == name:
                        status_name=False
                        status=False
                        break
            excelfile.close()
        if status == False:
            info=""
            if status_name == False:
                if info == "":
                    info += "Tên sản phẩm"
                else:
                    info += ", Tên sản phẩm"
            if status_unit == False:
                if info == "":
                    info += "Đơn vị"
                else:
                    info += ", Đơn vị"
            if status_quantity == False:
                if info == "":
                    info += "Số lượng"
                else:
                    info += ", Số lượng"
            if status_category == False:
                if info == "":
                    info += "Danh mục"
                else:
                    info += ", Danh mục"
            if status_price == False:
                if info == "":
                    info += "Giá bán"
                else:
                    info += ", Giá bán"
            if status_cost == False:
                if info == "":
                    info += "Giá nhập"
                else:
                    info += ", Giá nhập"
            if status_barcode == False:
                if info == "":
                    info += "Barcode"
                else:
                    info += ", Barcode"
            messagebox.showerror("Thông báo",info)
        else:
            self.SaveProductInfo(name,quantity,unit,category,price,cost,barcode,brand)
    def CreateBarcode(self):
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        status=True
        while True:
            barcode = random.choice(["A","B","C","D","E","F","G","H","I","K","L","M",'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'])
            barcode += random.choice(["A","B","C","D","E","F","G","H","I","K","L","M",'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'])
            barcode += str(random.randint(100000,999999))
            for cell in data["G"]:
                if barcode == cell.value:
                    status=False
                    break
                else:
                    pass
            if status == False:
                continue
            else:
                barcode_entry.delete(0, END)
                barcode_entry.insert(END, barcode)
                break
def OpenNewProductWindow():
    global NewProductWindow
    NewProductWindow=NewProduct()
# OpenNewProductWindow()