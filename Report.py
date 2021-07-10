import tkinter as tk
from tkinter import*
from openpyxl import load_workbook
from datetime import datetime,timedelta,date
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from pandas import DataFrame
class Report(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        global name_entry, barcode_entry,tree_importproducts,revenue_per_order,month_orders,month_revenue,today_revenue,chart1,chart2
        # create window
        self.geometry("1150x600+400+400")
        self.title("Báo cáo doanh thu, khách hàng")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self)
        bg = Label(self)
        bg.place(relx=0, rely=0, width=1150, height=600)
        bg_img = PhotoImage(master=canvas, width=1150, height=600, file="report.png")
        bg.configure(image=bg_img)

        today_revenue = StringVar(self)
        today_revenue_entry = Entry(self, font="Montserrat", state="readonly",
                                      textvariable=today_revenue,
                                      readonlybackground="#ffffff", cursor="hand2",relief="flat")
        today_revenue_entry.place(x=56, y=223, width=207, height=19)

        month_revenue = StringVar(self)
        month_revenue_entry = Entry(self, font="Montserrat", state="readonly",
                                    textvariable=month_revenue,
                                    readonlybackground="#ffffff", cursor="hand2", relief="flat")
        month_revenue_entry.place(x=56, y=304, width=207, height=19)

        month_orders = StringVar(self)
        month_orders_entry = Entry(self, font="Montserrat", state="readonly",
                                    textvariable=month_orders,
                                    readonlybackground="#ffffff", cursor="hand2", relief="flat")
        month_orders_entry.place(x=56, y=385, width=207, height=19)

        revenue_per_order = StringVar(self)
        revenue_per_order_entry = Entry(self, font="Montserrat", state="readonly",
                                    textvariable=revenue_per_order,
                                    readonlybackground="#ffffff", cursor="hand2", relief="flat")
        revenue_per_order_entry.place(x=56, y=466, width=207, height=19)

        chart1=Frame(self)
        chart1.place(x=331,y=162,w=350,h=353)

        chart2 = Frame(self)
        chart2.place(x=747, y=162, w=350, h=353)
        self.GeneralInformation()
        self.Charts()
        self.mainloop()
    def GeneralInformation(self):
        excelfile=load_workbook("TransactionManagement.xlsx")
        orders=0
        revenue=0
        today=0
        now = datetime.now()
        now = now.strftime("%d/%m/%Y %H").split(" ")
        now[0]=now[0].split("/")
        for row in excelfile.worksheets[0]:
            if row[3].value=="IN":
                order_time=row[4].value.split(" ")
                order_time[0]=order_time[0].split("/")
                if int(now[0][1])==int(order_time[0][1]):
                    orders+=1
                    if int(now[0][0])==int(order_time[0][0]):
                        try:
                            today+=int(row[2].value)
                        except:
                            today+=int(row[2].value.replace(",",""))
                    try:
                        revenue+=int(row[2].value)
                    except:
                        revenue+=int(row[2].value.replace(",",""))
                else:
                    pass
        revenue_per_order.set(str("{:,}".format(round(revenue/orders))))
        month_orders.set(str(orders))
        month_revenue.set(str("{:,}".format(revenue)))
        today_revenue.set(str("{:,}".format(today)))
        excelfile.close()
    def Charts(self):
        excelfile=load_workbook("TransactionManagement.xlsx")
        revenue0=0
        revenue1=0
        revenue2=0
        revenue3=0
        revenue4=0
        name0 = date.today()
        name1 = name0 - timedelta(days=1)
        name2 = name0 - timedelta(days=2)
        name3 = name0 - timedelta(days=3)
        name4 = name0 - timedelta(days=4)
        name0=str(name0)[8:10]
        name1 = str(name1)[8:10]
        name2 = str(name2)[8:10]
        name3 = str(name3)[8:10]
        name4 = str(name4)[8:10]
        now = datetime.now()
        now = now.strftime("%d/%m/%Y %H").split(" ")
        now[0] = now[0].split("/")
        start = datetime(year=int(now[0][2]), month=int(now[0][1]), day=int(now[0][0]))
        for row in excelfile.worksheets[0]:
            if row[3].value == "IN":
                order_time = row[4].value.split(" ")
                order_time[0] = order_time[0].split("/")
                end = datetime(year=int(order_time[0][2]), month=int(order_time[0][1]), day=int(order_time[0][0]))
                difference=start-end
                if timedelta(0,difference.total_seconds()).days==4:
                    try:
                        revenue4+=int(row[2].value)
                    except:
                        revenue4+=int(row[2].value.replace(",",""))
                elif timedelta(0,difference.total_seconds()).days==3:
                    try:
                        revenue3+=int(row[2].value)
                    except:
                        revenue3+=int(row[2].value.replace(",",""))
                elif timedelta(0,difference.total_seconds()).days==2:
                    try:
                        revenue2+=int(row[2].value)
                    except:
                        revenue2+=int(row[2].value.replace(",",""))
                elif timedelta(0,difference.total_seconds()).days==1:
                    try:
                        revenue1+=int(row[2].value)
                    except:
                        revenue1+=int(row[2].value.replace(",",""))
                elif timedelta(0,difference.total_seconds()).days==0:
                    try:
                        revenue0+=int(row[2].value)
                    except:
                        revenue0+=int(row[2].value.replace(",",""))
        excelfile.close()
        data1 = {'Doanh thu': [revenue4,revenue3,revenue2,revenue1,revenue0],
                 'Ngày': [name4,name3,name2,name1,name0]
                 }
        df1 = DataFrame(data1, columns=['Doanh thu', 'Ngày'])
        revenue = plt.Figure(figsize=(5, 2), dpi=115)
        ax1 = revenue.add_subplot(111)
        bar1 = FigureCanvasTkAgg(revenue, chart1)
        bar1.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH)
        df1 = df1[['Doanh thu', 'Ngày']].groupby('Ngày').sum()
        df1.plot(kind='bar', legend=True, ax=ax1)

        excelfile=load_workbook("CustomerManagement.xlsx")
        customer0=0
        customer1 = 0
        customer2 = 0
        customer3 = 0
        customer4 = 0
        for row in excelfile.worksheets[0]:
            registration_time = row[8].value.split(" ")
            registration_time[0] = registration_time[0].split("/")
            try:
                end = datetime(year=int(registration_time[0][2]), month=int(registration_time[0][1]), day=int(registration_time[0][0]))
                difference = start - end
                if timedelta(0, difference.total_seconds()).days == 4:
                    customer4 += 1
                elif timedelta(0, difference.total_seconds()).days == 3:
                    customer3 += 1
                elif timedelta(0, difference.total_seconds()).days == 2:
                    customer2 += 1
                elif timedelta(0, difference.total_seconds()).days == 1:
                    customer1 += 1
                elif timedelta(0, difference.total_seconds()).days == 0:
                    customer0 += 1
            except:
                continue
        excelfile.close()

        data2 = {'Khách hàng': [customer4,customer3,customer2,customer1,customer0],
                 'Ngày': [name4, name3, name2, name1, name0]
                 }
        df2 = DataFrame(data2, columns=['Khách hàng', 'Ngày'])
        figure2 = plt.Figure(figsize=(5, 3), dpi=100)
        ax2 = figure2.add_subplot(111)
        line2 = FigureCanvasTkAgg(figure2, chart2)
        line2.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH)
        df2 = df2[['Khách hàng', 'Ngày']].groupby('Ngày').sum()
        df2.plot(kind='line', legend=True, ax=ax2, color='r', marker='o', fontsize=10)
Report()