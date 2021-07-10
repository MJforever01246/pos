import tkinter as tk
from tkinter import*
from tkinter import ttk
from tkinter import messagebox as mb
from openpyxl import load_workbook
from FullScreen import FullScreenApp
from datetime import datetime
class CustomerManagement(tk.Tk):
    def __init__(self):
        global tree
        # create self
        tk.Tk.__init__(self)
        FullScreenApp(self)
        self.title("Quản lí khách hàng")
        # canvas: background and buttons
        canvas=Canvas(master=self,width=1920,height=1080,bg="#000000")
        bg=Label(master=self)
        bg_img=PhotoImage(file="CustomerManagement.png",master=canvas)
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0,width=1920,height=1080)
        Search = Button(self, command=lambda: self.OpenSearchWindow("search"))
        Search.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        search_img = PhotoImage(file="SearchButton.png", master=canvas)
        Search.configure(image=search_img)
        Search.place(x=62, y=400)
        AddCustomer = Button(self)
        AddCustomer.configure(command=lambda:AddCustomerWindow("Admin"),relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                              foreground="#ffffff", background="#ffffff", borderwidth="0")
        AddCustomer_img = PhotoImage(file="AddCustomer.png", master=canvas)
        AddCustomer.configure(image=AddCustomer_img)
        AddCustomer.place(x=62, y=500)
        Edit = Button(self, command=lambda: self.OpenSearchWindow("Edit"))
        Edit.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                       foreground="#ffffff", background="#ffffff", borderwidth="0")
        Edit_img = PhotoImage(file="EditButton.png", master=canvas)
        Edit.configure(image=Edit_img)
        Edit.place(x=62, y=600)
        Delete = Button(self, command=lambda: self.OpenSearchWindow("Delete"))
        Delete.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        Delete_img = PhotoImage(file="DeleteAccount.png", master=canvas)
        Delete.configure(image=Delete_img)
        Delete.place(x=62, y=700)
        # table
        TableMargin = Frame(self, width=500)
        TableMargin.place(x=350, y=90)
        scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
        scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
        tree = ttk.Treeview(TableMargin,
                            columns=("Name", "DOB","Gender","Email", "Phone number", "Reward","Loyalty program","Total spending",'Registration date'),
                            height=42, selectmode="extended", yscrollcommand=scrollbary.set,
                            xscrollcommand=scrollbarx.set)
        scrollbary.config(command=tree.yview)
        scrollbary.pack(side=RIGHT, fill=Y)
        scrollbarx.config(command=tree.xview)
        scrollbarx.pack(side=BOTTOM, fill=X)
        tree.heading('Name', text="Tên")
        tree.heading('DOB', text="DOB")
        tree.heading('Gender', text="Giới tính")
        tree.heading('Email', text="Email")
        tree.heading('Phone number', text="Di động")
        tree.heading('Reward', text="Điểm thưởng")
        tree.heading('Loyalty program',text="Nhóm khách hàng")
        tree.heading("Total spending",text="Tổng chi tiêu")
        tree.heading('Registration date', text="Ngày đăng kí")
        tree.column('#0', stretch=NO, minwidth=0, width=0)
        tree.column('#1', stretch=NO, minwidth=0, width=160)
        tree.column('#2', stretch=NO, minwidth=0, width=160)
        tree.column('#3', stretch=NO, minwidth=0, width=160)
        tree.column('#4', stretch=NO, minwidth=0, width=160)
        tree.column('#5', stretch=NO, minwidth=0, width=160)
        tree.column('#6', stretch=NO, minwidth=0, width=160)
        tree.column('#7', stretch=NO, minwidth=0, width=160)
        tree.column('#8', stretch=NO, minwidth=0, width=160)
        tree.column('#9', stretch=NO, minwidth=0, width=160)
        tree.pack()
        excelfile = load_workbook("CustomerManagement.xlsx")
        for row in excelfile.worksheets[0]:
            if row[0].value == "Name" and row[1].value == "DOB":
                pass
            else:
                tree.insert("", 0, values=(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value,row[6].value,row[7].value,row[8].value))
        excelfile.close()
        # Search
        self.mainloop()
    def OpenSearchWindow(self,reason):
        global window_search
        window_search=SearchWindow(reason)
class SearchWindow(tk.Tk):
    def __init__(self,reason):
        global Email_Phone, window_search
        tk.Tk.__init__(self)
        self.geometry("400x200+400+400")
        self.title("Tìm kiếm khách hàng")
        self.resizable(0, 0)
        back_search = Frame(master=self, bg='white')
        back_search.pack_propagate(0)
        back_search.pack(expand=1)
        canvas1=Canvas(master=self, width=400, height=200, bg="#000000")
        bg_search = Label(master=self)
        bg_search.place(relx=0, rely=0, width=400, height=200)
        bg_img = PhotoImage(master=canvas1, width=400, height=200, file="SearchCustomer_Employee.png")
        bg_search.configure(image=bg_img)
        Email_Phone = Entry(master=self,relief="flat")
        Email_Phone.place(x=80, y=90,height=24,width=290)
        Search = Button(self,command=lambda: self.SearchWindowProcess(reason))
        Search.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        Search_img = PhotoImage(file="Search.png", master=canvas1)
        Search.configure(image=Search_img)
        Search.place(x=135, y=150)
        self.mainloop()
    def SearchWindowProcess(self,reason): #There are three cases which use this function: Edit, Register, Delete Account
        global list_profile,i
        Email_Phone_info = Email_Phone.get()
        self.destroy()
        excelfile=load_workbook("CustomerManagement.xlsx")
        status = ""
        OK = False
        i = -2
        if "@" in Email_Phone_info or "." in Email_Phone_info:
            status = "Email"
        for row in excelfile.worksheets[0]:
            i += 1
            if status == "Email":
                if row[3].value == Email_Phone_info:
                    OK = True
                    if reason == "Edit":
                        list_profile = []
                        for w in range(0, 7):
                            if w == 5:
                                pass
                            else:
                                list_profile.append(row[w].value)
                    break
            else:
                if row[4].value == Email_Phone_info:
                    OK = True
                    if reason == "Edit":
                        list_profile = []
                        for w in range(0, 7):
                            if w == 5:
                                pass
                            else:
                                list_profile.append(row[w].value)
                    break
        if OK == False:
            mb.showerror("Thông báo","Không tìm thấy thông tin")
        else:
            if reason == "Delete":
                self.remove_acc()
                mb.showinfo("Thông báo","Xóa tài khoản thành công")
            else:
                mb.showinfo("Thông báo", "Tìm kiếm thông tin thành công")
            if reason == "Edit":
                EditCustomerWindow("Admin")
            else:
                tree.selection_set(tree.get_children("")[-i-1])
                OK=False
        excelfile.close()
    def remove_acc(self):
        excelfile=load_workbook("CustomerManagement.xlsx")
        data=excelfile["CustomerManagement"]
        data.delete_rows(idx=i+2,amount=1)
        excelfile.save("CustomerManagement.xlsx")
        excelfile.close()
        selected_item=tree.get_children("")[-i-1]
        tree.delete(selected_item)
class EditCustomerWindow(tk.Tk):
    def __init__(self,account_type):
        tk.Tk.__init__(self)
        global name_entry, dob, option_gender, email_entry, phone_entry,loyaltyprogram_entry,value_tuple,order
        # create window
        self.geometry("1150x600+400+400")
        self.title("Sửa thông tin khách hàng")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="editcustomer.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        # name
        name_entry = Entry(self, width=55,relief="flat")
        name_entry.place(x=90, y=197)
        name_entry.insert(END,list_profile[0])
        # DOB
        dob = Entry(self,width=55,relief="flat")
        dob.place(x=90, y=286)
        dob.insert(END,list_profile[1])
        # Gender
        option_gender=StringVar(master=self)
        option_gender.set(list_profile[2])
        img_male=PhotoImage(master=canvas,file="RadioButtonMale.png")
        CheckButtonMale=Checkbutton(master=self,image=img_male,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nam",command=lambda:self.Checkboxgender("Nam"),offvalue="")
        CheckButtonMale.place(x=144, y=371)
        img_female=PhotoImage(master=canvas,file="RadioButtonFemale.png")
        CheckButtonFemale=Checkbutton(master=self,image=img_female,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nữ",command=lambda:self.Checkboxgender("Nữ"),offvalue="")
        CheckButtonFemale.place(x=280, y=371)
        # email
        email_entry = Entry(self, width=55,relief="flat")
        email_entry.place(x=708, y=197)
        email_entry.insert(END,list_profile[3])
        # phone
        phone_entry = Entry(self, width=55,relief="flat")
        phone_entry.place(x=708, y=286)
        phone_entry.insert(END,list_profile[4])
        # loyalty program
        loyaltyprogram=StringVar(master=self)
        loyaltyprogram_entry = Spinbox(self, width=54,relief="flat")
        loyaltyprogram_entry.place(x=708, y=375)
        if account_type=="Thu ngân":
            loyaltyprogram_entry.configure(value=(str(list_profile[5])),state="disabled",disabledbackground="#ffffff")
        else:
            value_tuple=[]
            excelfile=load_workbook("CustomerManagement.xlsx")
            for row in excelfile.worksheets[1]:
                if row[0].value=="Name":
                    pass
                else:
                    value_tuple.append(row[0].value)
            value_tuple=tuple(value_tuple)
            loyaltyprogram_entry.configure(value=value_tuple,textvariable=loyaltyprogram)
            loyaltyprogram.set(list_profile[5])
            excelfile.close()
        # button
        complete_img=PhotoImage(master=canvas,file="Save.png")
        Complete = Button(self,command=self.process,image=complete_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        Complete.place(x=505, y=550)
        order=i
        self.mainloop()
    def Checkboxgender(self,option):
        option_gender.set(option)
    def process(self):
        global email_info, phone_info, status, otp, window_otp,gender
        status = True
        status_name=True
        email_info = email_entry.get()
        phone_info = phone_entry.get()
        status_email = True
        status_phone = True
        status_loyaltyprogram=True
        status_dob = True
        excelfile = load_workbook("CustomerManagement.xlsx")
        customer_data=excelfile["CustomerManagement"]
        if list_profile[3] == email_info:
            pass
        else:
            for cell in customer_data["D"]:
                if cell.value == list_profile[3]:
                    pass
                elif cell.value == email_info:
                    status = False
                    status_email = False
                    break
        if list_profile[4] == phone_info:
            pass
        else:
            for cell in customer_data["E"]:
                if cell.value == list_profile[4]:
                    pass
                elif cell.value == phone_info:
                    status_phone= False
                    status = False
                    break
        excelfile.close()
        if name_entry.get()=="":
            status_name=False
            status=False
        if status_email == True:
            if email_info.find("@") == -1 or email_info.find(".") == -1:
                status = False
                status_email = False
        for i in phone_info:
            if i in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                pass
            else:
                status_phone = False
                status = False
                break
        if status_phone == True:
            if len(phone_info) == 10 and phone_info[0] == "0":
                pass
            else:
                status = False
                status_phone = False
        if loyaltyprogram_entry.get() in value_tuple:
            pass
        else:
            status=False
            status_loyaltyprogram=False
        if dob.get()=="":
            pass
        else:
            cache=dob.get()
            cache=cache.split("/")
            try:
                if int(cache[0]) > 0 and int(cache[0]) < 32 and int(cache[1]) > 0 and int(cache[1]) <13 and int(cache[2]) > 1900 and int(cache[2]) < 2021:
                    pass
                else:
                    status=False
                    status_dob=False
            except:
                status = False
                status_dob = False
        if status == False:
            info = ""
            if status_name==False:
                if info == "":
                    info += "Tên"
                else:
                    info += ", Tên"
            if status_loyaltyprogram == False:
                if info == "":
                    info += "Nhóm khách hàng"
                else:
                    info += ", Nhóm khách hàng"
            if status_dob == False:
                if info == "":
                    info += "Ngày sinh"
                else:
                    info += ", Ngày sinh"
            if status_phone == False:
                if info == "":
                    info += "Di động"
                else:
                    info += ", Di động"
            if status_email == False:
                if info == "":
                    info += "Email"
                else:
                    info += ", Email"
            mb.showerror("Thông báo", info + " không hợp lệ",master=self)
        if status == True:
            self.SaveCustomerInfo(name_entry.get(),dob.get(),option_gender.get(),email_info,phone_info,loyaltyprogram_entry.get())
    def SaveCustomerInfo(self,name,dob,gender,email,phone,loyaltyprogram):
        excelfile = load_workbook("CustomerManagement.xlsx")
        info=[]
        for row in excelfile.worksheets[0]:
            if row[4].value == list_profile[4]:
                row[0].value = str(name)
                row[1].value = str(dob)
                row[2].value = str(gender)
                row[3].value = str(email)
                row[4].value = str(phone)
                row[6].value = str(loyaltyprogram)
                info.append(row[5].value)
                info.append(row[7].value)
                info.append(row[8].value)
                excelfile.save("CustomerManagement.xlsx")
                excelfile.close()
                break
        selected_item = tree.get_children("")[-order - 1]
        tree.delete(selected_item)
        tree.insert("", -order-1, values=(name,dob,gender,email,phone,info[0],loyaltyprogram,info[1],info[2]))
        self.destroy()
        mb.showinfo("Thông báo", "Thay đổi thông tin thành công")
class AddCustomerWindow(tk.Tk):
    def __init__(self,account_type):
        tk.Tk.__init__(self)
        global name_entry, dob, option_gender, email_entry, phone_entry,loyaltyprogram_entry,value_tuple
        # create window
        self.geometry("1150x600+400+400")
        self.title("Thêm khách hàng")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="addcustomer_bg.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        # name
        name_entry = Entry(self, width=55,relief="flat")
        name_entry.place(x=90, y=197)
        # DOB
        dob = Entry(self,width=55,relief="flat")
        dob.place(x=90, y=286)
        # Gender
        option_gender=StringVar(master=self)
        option_gender.set("")
        img_male=PhotoImage(master=canvas,file="RadioButtonMale.png")
        CheckButtonMale=Checkbutton(master=self,image=img_male,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nam",command=lambda:self.Checkboxgender("Nam"),offvalue="")
        CheckButtonMale.place(x=144, y=371)
        img_female=PhotoImage(master=canvas,file="RadioButtonFemale.png")
        CheckButtonFemale=Checkbutton(master=self,image=img_female,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nữ",command=lambda:self.Checkboxgender("Nữ"),offvalue="")
        CheckButtonFemale.place(x=280, y=371)
        # email
        email_entry = Entry(self, width=55,relief="flat")
        email_entry.place(x=708, y=197)
        # phone
        phone_entry = Entry(self, width=55,relief="flat")
        phone_entry.place(x=708, y=286)
        # loyalty program
        loyaltyprogram_entry = Spinbox(self, width=54,relief="flat")
        loyaltyprogram_entry.place(x=708, y=375)
        if account_type=="Thu ngân":
            excelfile = load_workbook("CustomerManagement.xlsx")
            for row in excelfile.worksheets[1]:
                if row[0].value == "Name":
                    pass
                else:
                    if row[1].value==0:
                        loyaltyprogram_entry.configure(value=(str(row[0].value)),state="disabled",disabledbackground="#ffffff")
                    else:
                        continue
        else:
            value_tuple=[]
            excelfile=load_workbook("CustomerManagement.xlsx")
            for row in excelfile.worksheets[1]:
                if row[0].value=="Name":
                    pass
                else:
                    value_tuple.append(row[0].value)
            value_tuple=tuple(value_tuple)
            loyaltyprogram_entry.configure(value=value_tuple)
            excelfile.close()
        # button
        complete_img=PhotoImage(master=canvas,file="Save.png")
        Complete = Button(self,command=self.process,image=complete_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        Complete.place(x=505, y=550)
        self.mainloop()
    def Checkboxgender(self,option):
        option_gender.set(option)
    def process(self):
        global email_info, phone_info, status, otp, window_otp,gender
        status = True
        status_name=True
        email_info = email_entry.get()
        phone_info = phone_entry.get()
        status_email = True
        status_phone = True
        status_loyaltyprogram=True
        status_dob = True
        excelfile = load_workbook("CustomerManagement.xlsx")
        for row in excelfile.worksheets[0]:
            if email_info=="":
                status_email=None
            else:
                if row[3].value == email_info:
                    status = False
                    status_email = False
            if row[4].value == phone_info:
                status = False
                status_phone = False
        excelfile.close()
        if name_entry.get()=="":
            status_name=False
            status=False
        if status_email == True:
            if email_info.find("@") == -1 or email_info.find(".") == -1:
                status = False
                status_email = False
        for i in phone_info:
            if i in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                pass
            else:
                status_phone = False
                status = False
                break
        if status_phone == True:
            if len(phone_info) == 10 and phone_info[0] == "0":
                pass
            else:
                status = False
                status_phone = False
        if loyaltyprogram_entry.get() in value_tuple:
            pass
        else:
            status=False
            status_loyaltyprogram=False
        if dob.get()=="":
            pass
        else:
            cache=dob.get()
            cache=cache.split("/")
            try:
                if int(cache[0]) > 0 and int(cache[0]) < 32 and int(cache[1]) > 0 and int(cache[1]) <13 and int(cache[2]) > 1900 and int(cache[2]) < 2021:
                    pass
                else:
                    status=False
                    status_dob=False
            except:
                status = False
                status_dob = False
        if status == False:
            info = ""
            if status_name==False:
                if info == "":
                    info += "Tên"
                else:
                    info += ", Tên"
            if status_loyaltyprogram==False:
                if info == "":
                    info += "Nhóm khách hàng"
                else:
                    info += ", Nhóm khách hàng"
            if status_dob == False:
                if info == "":
                    info += "Ngày sinh"
                else:
                    info += ", Ngày sinh"
            if status_phone == False:
                if info == "":
                    info += "Di động"
                else:
                    info += ", Di động"
            if status_email == False:
                if info == "":
                    info += "Email"
                else:
                    info += ", Email"
            mb.showerror("Thông báo", info + " không hợp lệ",master=self)
        if status == True:
            self.SaveCustomerInfo(name_entry.get(),dob.get(),option_gender.get(),email_info,phone_info,loyaltyprogram_entry.get())
    def SaveCustomerInfo(self,name,dob,gender,email,phone,loyaltyprogram):
        excelfile = load_workbook("CustomerManagement.xlsx")
        data=excelfile["CustomerManagement"]
        time=datetime.now()
        time = time.strftime("%d/%m/%Y %H:%M:%S")
        savedata = [name,dob,gender,email,phone,0,loyaltyprogram,0,time]
        data.insert_rows(idx=2, amount=1)
        list_order = ["A2", "B2", "C2", "D2", "E2","F2", "G2","H2", "I2"]
        for i in range(9):
            data[list_order[i]] = savedata[i]
        excelfile.save("CustomerManagement.xlsx")
        savedata=tuple(savedata)
        tree.insert("", 0,values=savedata)
        mb.showinfo("Thông báo","Thêm khách hàng thành công",master=self)
        self.destroy()
        excelfile.close()
# window_management=CustomerManagement()