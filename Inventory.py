import tkinter as tk
from tkinter import*
from tkinter import ttk
from tkinter import messagebox as mb
# from AddProduct import OpenNewProductWindow
import random
from openpyxl import load_workbook
from FullScreen import FullScreenApp
from datetime import datetime
class Inventory(tk.Tk):
    def __init__(self):
        global tree
        # create self
        tk.Tk.__init__(self)
        FullScreenApp(self)
        self.title("Quản lí kho & sản phẩm")
        # title
        canvas = Canvas(self, width=1920, height=1080, bg="#000000")
        bg=Label(self)
        bg.place(relx=0,rely=0,width=1920,height=1080)
        bg_img=PhotoImage(master=canvas,width=1920,height=1080,file="InventoryBG.png")
        bg.configure(image=bg_img)
        # table
        TableMargin = Frame(self, width=500)
        TableMargin.place(x=350, y=80)
        scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
        scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
        tree = ttk.Treeview(TableMargin,
                            columns=("Name", "Category","Quantity","Unit", "Price", "Cost", "Barcode","Brand"),
                            height=42, selectmode="extended", yscrollcommand=scrollbary.set,
                            xscrollcommand=scrollbarx.set)
        scrollbary.config(command=tree.yview)
        scrollbary.pack(side=RIGHT, fill=Y)
        scrollbarx.config(command=tree.xview)
        scrollbarx.pack(side=BOTTOM, fill=X)
        tree.heading('Name', text="Tên sản phẩm")
        tree.heading('Category', text="Danh mục")
        tree.heading('Unit', text="Đơn vị")
        tree.heading('Price', text="Đơn giá")
        tree.heading('Cost', text="Giá nhập")
        tree.heading('Barcode', text="Barcode")
        tree.heading('Quantity', text="Số lượng")
        tree.heading('Brand', text="Nhãn hiệu (NSX)")
        tree.column('#0', stretch=NO, minwidth=0, width=0)
        tree.column('#1', stretch=NO, minwidth=0, width=200)
        tree.column('#2', stretch=NO, minwidth=0, width=200)
        tree.column('#3', stretch=NO, minwidth=0, width=80)
        tree.column('#4', stretch=NO, minwidth=0, width=200)
        tree.column('#5', stretch=NO, minwidth=0, width=200)
        tree.column('#6', stretch=NO, minwidth=0, width=200)
        tree.column('#7', stretch=NO, minwidth=0, width=200)
        tree.column('#8', stretch=NO, minwidth=0, width=200)
        tree.pack()
        self.load_treeview()
        # Search
        Search = Button(self, command=lambda:self.OpenSearchWindow("search"))

        Search.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        search_img = PhotoImage(file="SearchButton.png", master=canvas)
        Search.configure(image=search_img)
        Search.place(x=62, y=300)
        AddProduct = Button(self, command=NewProduct)
        AddProduct.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        AddProduct_img = PhotoImage(file="AddProductButton.png", master=canvas)
        AddProduct.configure(image=AddProduct_img)
        AddProduct.place(x=62, y=400)
        Edit = Button(self, command=lambda: self.OpenSearchWindow("Edit"))
        Edit.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        Edit_img = PhotoImage(file="EditButton.png", master=canvas)
        Edit.configure(image=Edit_img)
        Edit.place(x=62, y=500)
        Delete = Button(self, command=lambda: self.OpenSearchWindow("Delete"))
        Delete.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        Delete_img = PhotoImage(file="DeleteProductButton.png", master=canvas)
        Delete.configure(image=Delete_img)
        Delete.place(x=62, y=600)
        ImportProduct = Button(self, command=ImportProductsOpen)
        ImportProduct.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                                foreground="#ffffff", background="#ffffff", borderwidth="0")
        ImportProduct_img = PhotoImage(file="ImportProductButton.png", master=canvas)
        ImportProduct.configure(image=ImportProduct_img)
        ImportProduct.place(x=62, y=700)
        self.mainloop()
    def OpenSearchWindow(self,reason):
        global window_search
        window_search=SearchWindow(reason)
    def load_treeview(self):
        excelfile = load_workbook("ProductsList.xlsx")
        for sheet in excelfile.worksheets:
            for row in sheet.rows:
                if row[0].value == "Name" and row[1].value == "Category":
                    pass
                else:
                    tree.insert("", 0, values=(
                    row[0].value, row[1].value, row[2].value, row[3].value, "{:,}".format(int(row[4].value)),
                    "{:,}".format(int(row[5].value)), row[6].value, row[7].value))
        excelfile.close()
class NewProduct(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        global name_entry, barcode_entry
        # create window
        self.geometry("1150x600+400+400")
        self.title("Thêm sản phẩm mới")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="addproduct.png",width=1150,height=600)
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0,width=1150,height=600)
        # name
        name_entry = Entry(self, width=56,relief="flat")
        name_entry.place(x=82.5, y=125)
        # Category
        category_entry = Entry(self, width=56,relief="flat")
        category_entry.place(x=701, y=391)
        # Unit
        unit_entry = Entry(self, width=56,relief="flat")
        unit_entry.place(x=82.5, y=213)
        # Quantity
        quantity_entry = Entry(self,width=56,relief="flat")
        quantity_entry.place(x=82.5,y=302)
        # Price
        price_entry = Entry(self, width=56,relief="flat")
        price_entry.place(x=701, y=213)
        # Cost
        cost_entry = Entry(self, width=56,relief="flat")
        cost_entry.place(x=701, y=302)
        # Barcode
        barcode_entry = Entry(self, width=47,relief="flat")
        barcode_entry.place(x=753, y=124)
        barcode_auto_create=Button(self,command=self.CreateBarcode,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        create_img=PhotoImage(master=canvas,file="Create.png")
        barcode_auto_create.configure(image=create_img)
        barcode_auto_create.place(x=698,y=121)
        # Brand
        brand_entry = Entry(self, width=56,relief="flat")
        brand_entry.place(x=82.5, y=391)
        # button
        Complete = Button(self,command=lambda:self.CheckValid(name_entry.get(),quantity_entry.get(),unit_entry.get(),category_entry.get(),price_entry.get(),cost_entry.get(),barcode_entry.get(),brand_entry.get()),relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        buttom_img=PhotoImage(master=canvas,file="Save.png")
        Complete.configure(image=buttom_img)
        Complete.place(x=500, y=528)
        self.mainloop()
    def SaveProductInfo(self,name,quantity,unit,category,price,cost,barcode,brand):
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        list_order=["A2","B2","C2","D2","E2","F2","G2","H2"]
        data_list=[name,category,int(quantity),unit,int(price),int(cost),barcode,brand]
        data.insert_rows(idx=2,amount=1)
        for i in range (8):
            data[list_order[i]].value=data_list[i]
        data["F2"].number_format = '#,##0'
        data["E2"].number_format = '#,##0'
        excelfile.save("ProductsList.xlsx")
        excelfile.close()
        mb.showinfo("Thông báo","Thêm sản phầm thành công")
        self.destroy()
    def CheckValid(self,name,quantity,unit,category,price,cost,barcode,brand):
        status=True
        status_name=True
        status_unit=True
        status_category=True
        status_quantity=True
        status_price=True
        status_cost=True
        status_barcode=True
        if name == "":
            status_name=False
            status=False
        if quantity.isdigit():
            pass
        else:
            status = False
            status_quantity=False
        if unit=="":
            status=False
            status_unit=False
        if category == "":
            status=False
            status_category=False
        if price.isdigit():
            pass
        else:
            status=False
            status_price=False
        if cost.isdigit():
            pass
        else:
            status=False
            status_cost=False
        if barcode == "":
            status = False
            status_barcode=False
        if status_barcode==True or status_name==True:
            excelfile=load_workbook("ProductsList.xlsx")
            data=excelfile["ProductsList"]
            if status_barcode==True:
                for cell in data["G"]:
                    if cell == barcode:
                        status_barcode=False
                        status=False
                        break
            if status_name==True:
                for cell in data["A"]:
                    if cell == name:
                        status_name=False
                        status=False
                        break
            excelfile.close()
        if status == False:
            info=""
            if status_name == False:
                if info == "":
                    info += "Tên sản phẩm"
                else:
                    info += ", Tên sản phẩm"
            if status_unit == False:
                if info == "":
                    info += "Đơn vị"
                else:
                    info += ", Đơn vị"
            if status_quantity == False:
                if info == "":
                    info += "Số lượng"
                else:
                    info += ", Số lượng"
            if status_category == False:
                if info == "":
                    info += "Danh mục"
                else:
                    info += ", Danh mục"
            if status_price == False:
                if info == "":
                    info += "Giá bán"
                else:
                    info += ", Giá bán"
            if status_cost == False:
                if info == "":
                    info += "Giá nhập"
                else:
                    info += ", Giá nhập"
            if status_barcode == False:
                if info == "":
                    info += "Barcode"
                else:
                    info += ", Barcode"
            mb.showerror("Thông báo",info)
        else:
            self.SaveProductInfo(name,quantity,unit,category,price,cost,barcode,brand)
    def CreateBarcode(self):
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        status=True
        while True:
            barcode = random.choice(["A","B","C","D","E","F","G","H","I","K","L","M",'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'])
            barcode += random.choice(["A","B","C","D","E","F","G","H","I","K","L","M",'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z'])
            barcode += str(random.randint(100000,999999))
            for cell in data["G"]:
                if barcode == cell.value:
                    status=False
                    break
                else:
                    pass
            if status == False:
                continue
            else:
                barcode_entry.delete(0, END)
                barcode_entry.insert(END, barcode)
                break
def ImportProductsOpen():
    global window_importproduct
    window_importproduct = ImportProducts()
class SearchWindow(tk.Tk):
    def __init__(self,reason):
        global product_entry
        tk.Tk.__init__(self)
        self.geometry("400x200+400+400")
        self.title("Tìm kiếm sản phẩm")
        self.resizable(0, 0)
        back_search = Frame(master=self, bg='white')
        back_search.pack_propagate(0)
        back_search.pack(expand=1)
        canvas1=Canvas(self, width=400, height=200, bg="#000000")
        bg_search = Label(self)
        bg_search.place(relx=0, rely=0, width=400, height=200)
        bg_img = PhotoImage(master=canvas1, width=400, height=200, file="SearchBG.png")
        bg_search.configure(image=bg_img)
        # Search_title = Label(master=self, text="Tìm kiếm sản phẩm", font=("Arial", 18, "bold")).place(x=85, y=0)
        # Search = Label(master=self, text="Nhập barcode hoặc tên", font=("Arial", 14, "underline")).place(x=20, y=50)
        product_entry = Entry(master=self)
        product_entry.configure(relief="flat")
        product_entry.place(x=80, y=90,height=24,width=290)
        Search = Button(self, command=lambda:self.SearchWindowProcess(reason))
        Search.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        Search_img = PhotoImage(file="Search.png", master=canvas1)
        Search.configure(image=Search_img)
        Search.place(x=135, y=150)
        self.mainloop()
    def SearchWindowProcess(self,reason): #There are three cases which use this function: Edit, Register, Delete Account
        global list_product_info,i
        product=product_entry.get()
        self.destroy()
        excelfile=load_workbook("ProductsList.xlsx")
        OK = False
        i = -2
        for sheet in excelfile.worksheets:
            for row in sheet.rows:
                i+=1
                if row[0].value == product or row[6].value == product:
                    OK = True
                    if reason == "Edit":
                        list_product_info=[]
                        for w in range(0,8):
                            list_product_info.append(row[w].value)
                    break
        if OK == False:
            mb.showerror("Thông báo","Không tìm thấy thông tin")
        else:
            if reason == "Delete":
                self.remove_product()
                mb.showinfo("Thông báo","Xóa sản phẩm thành công")
            elif reason == "Edit":
                index_product=i+2
                EditProduct(index_product)
            else:
                mb.showinfo("Thông báo", "Tìm kiếm thông tin thành công")
                tree.selection_set(tree.get_children("")[-i-1])
                OK=False
        excelfile.close()
    def remove_product(self):
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        data.delete_rows(idx=i+2,amount=1)
        excelfile.save("ProductsList.xlsx")
        excelfile.close()
        selected_item=tree.get_children("")[-i-1]
        tree.delete(selected_item)
class EditProduct(tk.Tk):
    def __init__(self,index_product):
        tk.Tk.__init__(self)
        global name_entry, barcode_entry,row_number
        # create window
        self.geometry("1150x600+400+400")
        self.title("Chỉnh sửa sản phẩm")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="editproduct.png",width=1150,height=600)
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0,width=1150,height=600)
        # name
        name_entry = Entry(self, width=56,relief="flat")
        name_entry.place(x=82.5, y=125)
        # Category
        category_entry = Entry(self, width=56,relief="flat")
        category_entry.place(x=701, y=391)
        # Unit
        unit_entry = Entry(self, width=56,relief="flat")
        unit_entry.place(x=82.5, y=213)
        # Quantity
        quantity_entry = Entry(self,width=56,relief="flat")
        quantity_entry.place(x=82.5,y=302)
        # Price
        price_entry = Entry(self, width=56,relief="flat")
        price_entry.place(x=701, y=213)
        # Cost
        cost_entry = Entry(self, width=56,relief="flat")
        cost_entry.place(x=701, y=302)
        # Barcode
        barcode_entry = Entry(self, width=47,relief="flat")
        barcode_entry.place(x=753, y=124)
        barcode_auto_create=Button(self,command=self.CreateBarcode,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        create_img=PhotoImage(master=canvas,file="Create.png")
        barcode_auto_create.configure(image=create_img)
        barcode_auto_create.place(x=698,y=121)
        # Brand
        brand_entry = Entry(self, width=56,relief="flat")
        brand_entry.place(x=82.5, y=391)
        # button
        Complete = Button(self,command=lambda:self.CheckValid(name_entry.get(),quantity_entry.get(),unit_entry.get(),category_entry.get(),price_entry.get(),cost_entry.get(),barcode_entry.get(),brand_entry.get()),relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        buttom_img=PhotoImage(master=canvas,file="Save.png")
        Complete.configure(image=buttom_img)
        Complete.place(x=500, y=528)
        name_entry.insert(END,list_product_info[0])
        category_entry.insert(END,list_product_info[1])
        unit_entry.insert(END,list_product_info[3])
        quantity_entry.insert(END,list_product_info[2])
        price_entry.insert(END,list_product_info[4])
        cost_entry.insert(END,list_product_info[5])
        barcode_entry.insert(END,list_product_info[6])
        brand_entry.insert(END,list_product_info[7])
        row_number=index_product
        self.mainloop()
    def SaveProductInfo(self, name, quantity, unit, category, price, cost, barcode, brand):
        excelfile = load_workbook("ProductsList.xlsx")
        data = excelfile["ProductsList"]
        data_list = [name, category, int(quantity), unit, int(price), int(cost), barcode, brand]
        list_order=["A2","B2","C2","D2","E2","F2","G2","H2"]
        data.delete_rows(idx=row_number,amount=1)
        data.insert_rows(idx=2, amount=1)
        for i in range(8):
            data[list_order[i]].value = data_list[i]
        data["F2"].number_format = '#,##0'
        data["E2"].number_format = '#,##0'
        excelfile.save("ProductsList.xlsx")
        excelfile.close()
        mb.showinfo("Thông báo", "Chỉnh sửa sản phầm thành công",master=self)
        self.destroy()
    def CheckValid(self, name, quantity, unit, category, price, cost, barcode, brand):
        status = True
        status_name = True
        status_unit = True
        status_category = True
        status_quantity = True
        status_price = True
        status_cost = True
        status_barcode = True
        if name == "":
            status_name = False
            status = False
        if quantity.isdigit():
            pass
        else:
            status = False
            status_quantity = False
        if unit == "":
            status = False
            status_unit = False
        if category == "":
            status = False
            status_category = False
        if price.isdigit():
            pass
        else:
            status = False
            status_price = False
        if cost.isdigit():
            pass
        else:
            status = False
            status_cost = False
        if barcode == "":
            status = False
            status_barcode = False
        if status_barcode==True or status_name==True:
            excelfile=load_workbook("ProductsList.xlsx")
            data=excelfile["ProductsList"]
            if status_barcode==True:
                for cell in data["G"]:
                    if cell == barcode:
                        status_barcode=False
                        status=False
                        break
            if status_name==True:
                for cell in data["A"]:
                    if cell == name:
                        status_name=False
                        status=False
                        break
            excelfile.close()
        if status == False:
            info = ""
            if status_name == False:
                if info == "":
                    info += "Tên sản phẩm"
                else:
                    info += ", Tên sản phẩm"
            if status_unit == False:
                if info == "":
                    info += "Đơn vị"
                else:
                    info += ", Đơn vị"
            if status_quantity == False:
                if info == "":
                    info += "Số lượng"
                else:
                    info += ", Số lượng"
            if status_category == False:
                if info == "":
                    info += "Danh mục"
                else:
                    info += ", Danh mục"
            if status_price == False:
                if info == "":
                    info += "Giá bán"
                else:
                    info += ", Giá bán"
            if status_cost == False:
                if info == "":
                    info += "Giá nhập"
                else:
                    info += ", Giá nhập"
            if status_barcode == False:
                if info == "":
                    info += "Barcode"
                else:
                    info += ", Barcode"
            mb.showerror("Thông báo", info,master=self)
        else:
            self.SaveProductInfo(name, quantity, unit, category, price, cost, barcode, brand)
            selected_item = tree.get_children("")[-i-1]
            tree.delete(selected_item)
            tree.insert("", -i-1, values=(name,category, quantity, unit, price, cost, barcode, brand))
    def CreateBarcode(self):
        excelfile = load_workbook("ProductsList.xlsx")
        data = excelfile["ProductsList"]
        status = True
        while True:
            barcode = random.choice(
                ["A", "B", "C", "D", "E", "F", "G", "H", "I", "K", "L", "M", 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                 'V', 'W', 'X', 'Y', 'Z'])
            barcode += random.choice(
                ["A", "B", "C", "D", "E", "F", "G", "H", "I", "K", "L", "M", 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                 'V', 'W', 'X', 'Y', 'Z'])
            barcode += str(random.randint(100000, 999999))
            for cell in data["G"]:
                if barcode == cell.value:
                    status = False
                    break
                else:
                    pass
            if status == False:
                continue
            else:
                barcode_entry.delete(0, END)
                barcode_entry.insert(END, barcode)
                break
class ImportProducts(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        global name_entry, barcode_entry,tree_importproducts
        # create window
        self.geometry("1150x600+400+400")
        self.title("Nhập kho")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        # title
        title = Label(master=self, text="Nhập kho", font=("Arial", 28, "bold"))
        title.place(x=500, y=0)
        # button
        canvas = Canvas(self, width=1150, height=600, bg="#000000")
        bg = Label(self)
        bg.place(relx=0, rely=0, width=1150, height=600)
        bg_img = PhotoImage(master = canvas, width = 1150, height = 600,file="ImportProducts.png")
        bg.configure(image=bg_img)
        CreateReceivingVoucher = Button(self,command=self.OpenReceivingVoucher)

        CreateReceivingVoucher.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        button_img=PhotoImage(file="CreateButton.png",master = canvas)
        CreateReceivingVoucher.configure(image=button_img)
        CreateReceivingVoucher.place(x=62, y=250)
        DeleteReceivingVoucher = Button(self,command=self.DeleteVoucher)
        DeleteReceivingVoucher.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                                         foreground="#ffffff", background="#ffffff", borderwidth="0")
        button_img2=PhotoImage(file="DeleteButton.png",master = canvas)
        DeleteReceivingVoucher.configure(image=button_img2)
        DeleteReceivingVoucher.place(x=62,y=350)
        # treeview
        TableMargin = Frame(self, width=500)
        TableMargin.place(x=350, y=80)
        scrollbarx = Scrollbar(TableMargin, orient=HORIZONTAL)
        scrollbary = Scrollbar(TableMargin, orient=VERTICAL)
        tree_importproducts = ttk.Treeview(TableMargin,
                            columns=("Reference","Barcode",	"Product","Quantity","Total amount","Time"),
                            height=22, selectmode="extended", yscrollcommand=scrollbary.set,
                            xscrollcommand=scrollbarx.set)
        scrollbary.config(command=tree_importproducts.yview)
        scrollbary.pack(side=RIGHT, fill=Y)
        scrollbarx.config(command=tree_importproducts.xview)
        scrollbarx.pack(side=BOTTOM, fill=X)
        tree_importproducts.heading('Reference', text="Mã phiếu")
        tree_importproducts.heading('Barcode', text="Barcode")
        tree_importproducts.heading("Product",text="Tên sản phẩm")
        tree_importproducts.heading('Quantity', text="Số lượng")
        tree_importproducts.heading('Total amount', text="Tổng tiền")
        tree_importproducts.heading('Time',text="Thời gian")
        tree_importproducts.column('#0', stretch=NO, minwidth=0, width=0)
        tree_importproducts.column('#1', stretch=NO, minwidth=0, width=125)
        tree_importproducts.column('#2', stretch=NO, minwidth=0, width=125)
        tree_importproducts.column('#3', stretch=NO, minwidth=0, width=125)
        tree_importproducts.column('#4', stretch=NO, minwidth=0, width=125)
        tree_importproducts.column('#5', stretch=NO, minwidth=0, width=125)
        tree_importproducts.column('#6', stretch=NO, minwidth=0, width=125)
        tree_importproducts.pack()
        excelfile = load_workbook("StockManagement.xlsx")
        for sheet in excelfile.worksheets:
            for row in sheet.rows:
                if row[0].value == "Reference" and row[1].value == "Barcode":
                    pass
                else:
                    tree_importproducts.insert("", 0, values=(
                    row[0].value, row[1].value, row[2].value, row[3].value, row[4].value,row[5].value))
        excelfile.close()
        self.mainloop()
    def OpenReceivingVoucher(self):
        window_receiving_voucher=ReceivingVoucher()
    def DeleteVoucher(self):
        messagebox=mb.askokcancel("Thông báo","Vui lòng chọn mục cần xóa ở bảng, sau đó bấm 'OK'\nĐể hủy, vui lòng bấm 'Cancel'")
        if messagebox==True:
            selected_item = tree_importproducts.focus()
            excelfile=load_workbook("StockManagement.xlsx")
            data=excelfile["StockManagement"]
            i=0
            for cell in data["A"]:
                i+=1
                if cell.value==tree_importproducts.item(selected_item)["values"][0]:
                    data.delete_rows(idx=i,amount=1)
                    break
            excelfile.save("StockManagement.xlsx")
            excelfile.close()
            i=0
            excelfile=load_workbook("TransactionManagement.xlsx")
            data=excelfile["TransactionManagement"]
            for cell in data["A"]:
                i+=1
                if cell.value==tree_importproducts.item(selected_item)["values"][0]:
                    data.delete_rows(idx=i,amount=1)
                    break
            excelfile.save("TransactionManagement.xlsx")
            excelfile.close()
            tree_importproducts.delete(selected_item)
class ReceivingVoucher(tk.Tk):
    def __init__(self):
        global reference_entry,barcode_entry,name_entry,quantity_entry,totalamount,quantity,name
        tk.Tk.__init__(self)
        self.geometry("1150x600+400+400")
        self.title("Tạo phiếu nhập kho")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="receivingvoucher.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0,width=1150,height=600)
        # Barcode
        barcode_entry = Entry(self, width=47,relief="flat")
        barcode_entry.place(x=136, y=166)
        search_button=Button(self,command=self.SearchProduct)
        search_button_img=PhotoImage(master=canvas,file="find.png")
        search_button.configure(image=search_button_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        search_button.place(x=81,y=164)
        # Name
        name=StringVar(master=self)
        name_entry = Entry(self, width=56,state="disabled",textvariable=name,relief="flat")
        name_entry.place(x=84, y=276)
        # Quantity
        quantity=StringVar(master=self)
        quantity_entry = Entry(self,textvariable=quantity, width=56,relief="flat")
        quantity_entry.place(x=702.5, y=276)
        reg_quantity=self.register(self.CalculateTotal)
        quantity_entry.config(validate="key",validatecommand=(reg_quantity,"%P"))
        # Reference
        reference_entry = Entry(self, width=47,relief="flat")
        reference_entry.place(x=754, y=166)
        reference_auto_create = Button(self,command=self.CreateReference)
        reference_img=PhotoImage(master=canvas,file="Create.png")
        reference_auto_create.configure(image=reference_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        reference_auto_create.place(x=699, y=164)
        # Total amount
        totalamount=StringVar(master=self)
        totalamount.set("0")
        totalamount_value=Entry(self,font=("Arial", 18),textvariable=totalamount,relief="flat")
        totalamount_value.place(x=215,y=400,width=210,height=50)
        #Button
        complete=Button(self,command=self.CheckValid)
        complete_img=PhotoImage(master=canvas,file="Save.png")
        complete.configure(image=complete_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        complete.place(x=800,y=395,height=50)
        self.mainloop()
    def CreateReference(self):
        excelfile = load_workbook("StockManagement.xlsx")
        data = excelfile["StockManagement"]
        status = True
        while True:
            reference = "IM"
            reference += str(random.randint(100000, 999999))
            for cell in data["A"]:
                if reference == cell.value:
                    status = False
                    break
                else:
                    pass
            if status == False:
                continue
            else:
                reference_entry.delete(0, END)
                reference_entry.insert(END, reference)
                break
    def SearchProduct(self):
        global cost,index,name
        barcode=barcode_entry.get()
        excelfile = load_workbook("ProductsList.xlsx")
        OK = False
        i = -2
        for sheet in excelfile.worksheets:
            for row in sheet.rows:
                i += 1
                if row[6].value == barcode:
                    OK = True
                    name.set(row[0].value)
                    cost=str(row[5].value)
                    break
        index=i
        if OK == False:
            mb.showerror("Thông báo", "Không tìm thấy thông tin\nNếu sản phẩm này mới, vui lòng thêm sản phẩm trước",master=self)
        excelfile.close()
    def CalculateTotal(self,input):
        global totalamount_value
        try:
            totalamount_value=eval(input+"*"+cost)
            totalamount.set(f"{totalamount_value:,}")
        except:
            totalamount.set("0")
        return True
    def CheckValid(self):
        status = True
        status_barcode = True
        status_reference = True
        status_quantity = True
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        for cell in data["G"]:
            if cell.value == barcode_entry.get():
                status=True
                status_barcode=True
                break
            status=False
            status_barcode=False
        excelfile.close()
        excelfile=load_workbook("StockManagement.xlsx")
        data=excelfile["StockManagement"]
        for cell in data["A"]:
            if cell.value == reference_entry:
                status=False
                status_reference=False
                break
        excelfile.close()
        value=quantity.get()
        if value.isdigit():
            pass
        else:
            status=False
            status_reference=False
        if reference_entry.get()=="":
            status=False
            status_quantity=False
        if status == False:
            info = ""
            if status_quantity == False:
                if info=="":
                    info+="Số lượng không hợp lệ"
                else:
                    info+="; Số lượng không hợp lệ"
            if status_reference == False:
                if info=="":
                    info+="Mã phiếu đã tồn tại"
                else:
                    info+="; Mã phiếu đã tồn tại"
            if status_barcode == False:
                if info=="":
                    info+="Barcode không tồn tại"
                else:
                    info+="; Barcode không tồn tại"
            mb.showerror("Thông báo", info,master=self)
        else:
            self.SaveStockManagement(reference_entry.get(),barcode_entry.get(),name_entry.get(),quantity.get(),totalamount.get())
    def SaveStockManagement(self,Reference,Barcode,Product,Quantity,Totalamount):
        excelfile=load_workbook("StockManagement.xlsx")
        data=excelfile["StockManagement"]
        data.insert_rows(idx=2,amount=1)
        time=datetime.now()
        time=time.strftime("%d/%m/%Y %H:%M:%S")
        list_order = ["A2", "B2", "C2", "D2", "E2","F2"]
        list_info=[Reference,Barcode,Product,int(Quantity),Totalamount,time]
        for i in range (6):
            data[list_order[i]].value=list_info[i]
        data["E2"].number_format = '#,##0'
        excelfile.save("StockManagement.xlsx")
        excelfile.close()
        excelfile=load_workbook("ProductsList.xlsx")
        data=excelfile["ProductsList"]
        data["C"+str(index+2)].value=int(data["C"+str(index+2)].value)+int(Quantity)
        info=[]
        for i in ["A","B","C","D","E","F","G","H"]:
            info.append(data[i+str(index+2)].value)
        excelfile.save("ProductsList.xlsx")
        excelfile.close()
        selected_item = tree.get_children("")[-index-1]
        tree.delete(selected_item)
        tree.insert("", 1, values=(tuple(info)))
        excelfile=load_workbook("TransactionManagement.xlsx")
        data=excelfile["TransactionManagement"]
        data.insert_rows(idx=2,amount=1)
        data["A2"].value=Reference
        data["B2"].value="Import: "+str(Product)
        data["C2"].value=Totalamount
        data["C2"].number_format =  '#,##0'
        data["D2"].value="OUT"
        data["E2"].value=time
        excelfile.save("TransactionManagement.xlsx")
        excelfile.close()
        list_info=tuple(list_info)
        tree_importproducts.insert("", 0, values=(list_info))
        mb.showinfo("Thông báo","Nhập kho thành công",master=self)
        self.destroy()
def EmployeesManagementOpen():
    global window_management
    window_management=Inventory()
# EmployeesManagementOpen()
# ImportProductsOpen()