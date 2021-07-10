from tkinter import *
import tkinter as tk
from tkinter import messagebox as mb
from openpyxl import load_workbook,Workbook
import hashlib
from Mainsession import MainSession
from OTP import GenerateOTP
from Dashboard import Dashboard
import os
from EmployeesManagement import RegisterWindowForUser
def InitialSetup():
    if os.path.isfile('./CustomerManagement.xlsx'):
        pass
    else:
        excelfile=Workbook()
        sheet=excelfile.worksheets[0]
        sheet.title="CustomerManagement"
        headings=["Name","DOB","Gender","Email","Phone number","Reward","Loyalty program","Total spending","Registration date"]
        cells=["A1","B1","C1","D1","E1","F1","G1","H1","I1"]
        order=0
        for i in cells:
            sheet[i].value=headings[order]
            order+=1
        excelfile.save("CustomerManagement.xlsx")
        excelfile.close()
    if os.path.isfile('./ProductsList.xlsx'):
        pass
    else:
        excelfile=Workbook()
        sheet=excelfile.worksheets[0]
        sheet.title="ProductsList"
        headings=["Name","Category","In stock","Sold by (Weight/Volume/Each)","Price","Cost","Barcode","Brand"]
        cells=["A1","B1","C1","D1","E1","F1","G1","H1"]
        order=0
        for i in cells:
            sheet[i].value=headings[order]
            order+=1
        excelfile.save("ProductsList.xlsx")
        excelfile.close()
    if os.path.isfile('./StockManagement.xlsx'):
        pass
    else:
        excelfile=Workbook()
        sheet=excelfile.worksheets[0]
        sheet.title="StockManagement"
        headings=["Reference","Barcode","Product","Quantity","Total amount","Date"]
        cells=["A1","B1","C1","D1","E1","F1"]
        order=0
        for i in cells:
            sheet[i].value=headings[order]
            order+=1
        excelfile.save("StockManagement.xlsx")
        excelfile.close()
    if os.path.isfile('./TransactionManagement.xlsx'):
        pass
    else:
        excelfile = Workbook()
        sheet = excelfile.worksheets[0]
        sheet.title = "TransactionManagement"
        headings = ["Reference","Details","Amount","Cash In/Out","Date","Customer"]
        cells = ["A1", "B1", "C1", "D1", "E1", "F1"]
        order = 0
        for i in cells:
            sheet[i].value = headings[order]
            order += 1
        excelfile.save("TransactionManagement.xlsx")
        excelfile.close()
    if os.path.isfile('./UsersData.xlsx'):
        pass
    else:
        excelfile = Workbook()
        sheet = excelfile.worksheets[0]
        sheet.title = "UsersData"
        headings = ["Username","Password","Name","DOB","Gender","Email","Phone","Position"]
        cells = ["A1", "B1", "C1", "D1", "E1", "F1","G1","H1"]
        order = 0
        for i in cells:
            sheet[i].value = headings[order]
            order += 1
        excelfile.save("UsersData.xlsx")
        excelfile.close()
    if os.path.isfile('./Profile_store.xlsx'):
        pass
    else:
        excelfile=Workbook()
        sheet=excelfile.worksheets[0]
        headings=["Name","Address","Contact"]
        cells=["A1","B1","C1"]
        order=0
        for i in cells:
            sheet[i].value=headings[order]
            order+=1
        excelfile.save("Profile_store.xlsx")
        excelfile.close()
        RegisterWindowForUser()
        Setup()
class Setup(tk.Tk):
    def __init__(self):
        global address_entry,name_entry, phone_entry
        tk.Tk.__init__(self)
        self.geometry("400x400")
        self.resizable(0, 0)
        self.title("Thêm thông tin cửa hàng")
        canvas=Canvas(master=self,width=400,height=200)
        bg_img=PhotoImage(master=canvas,file="setup.png")
        bg=Label(self,image=bg_img)
        bg.place(x=0,y=0)
        button_img=PhotoImage(master=canvas,file="Save_delivery.png")
        save_button=Button(master=self,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0",command=self.CheckValid)
        save_button.configure(image=button_img)
        save_button.place(x=134,y=342)
        name_entry=Entry(self,font="Montserrat",relief="flat")
        name_entry.place(x=32,y=91,width=346,height=25)
        address_entry=Entry(self,font="Montserrat",relief="flat")
        address_entry.place(x=32,y=185,width=346,height=25)
        phone_entry_shipping = Entry(self, font="Montserrat", relief="flat")
        phone_entry_shipping.place(x=32, y=279, width=346, height=25)
        self.mainloop()
    def CheckValid(self):
        status=True
        status_address=True
        status_phone=True
        if len(phone_entry.get()) == 10 and phone_entry.get()[0] == "0":
            for i in phone_entry.get():
                if i in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                    pass
                else:
                    status_phone=False
                    status=False
        else:
            status_phone=False
            status=False
        if len(address_entry.get()) <=5:
            status_address=False
            status=False
        if status == False:
            info = ""
            if status_address == False:
                if info == "":
                    info += "Địa chỉ quá ngắn"
                else:
                    info += "\nĐịa chỉ quá ngắn"
            if status_phone == False:
                if info == "":
                    info += "Số điện thoại không hợp lệ"
                else:
                    info += "\nSố điện thoại không hợp lệ"
            mb.showerror("Thông báo",info,master=self)
        else:
            excelfile=load_workbook("Profile_store.xlsx")
            sheet = excelfile.worksheets[0]
            headings = [name_entry.get(), address_entry.get(), phone_entry.get()]
            cells = ["A2", "B2", "C2"]
            order = 0
            for i in cells:
                sheet[i].value = headings[order]
                order += 1
            excelfile.save("Profile_store.xlsx")
            excelfile.close()
            self.destroy()
class Log_in(tk.Tk):
    def __init__(self):
        global username_entry,password_entry
        tk.Tk.__init__(self)
        self.geometry("400x200")
        self.resizable(0, 0)
        self.title("Đăng nhập")
        canvas=Canvas(master=self,width=400,height=200)
        bg_img=PhotoImage(master=canvas,file="login.png")
        bg=Label(self,image=bg_img)
        bg.place(x=0,y=0)
        button_img=PhotoImage(master=canvas,file="OTPCheck.png")
        save_button=Button(master=self,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0",command=self.CheckValid)
        save_button.configure(image=button_img)
        save_button.place(x=134,y=150)
        username_entry=Entry(self,font="Montserrat",relief="flat")
        username_entry.place(x=97,y=62,width=283,height=25)
        password_entry = Entry(self, font="Montserrat", relief="flat",show="-")
        password_entry.place(x=97, y=99, width=283, height=25)
        self.mainloop()
    def CheckValid(self):
        status=False
        excelfile=load_workbook("UsersData.xlsx")
        password = password_entry.get()
        for row in excelfile.worksheets[0]:
            if row[0].value==username_entry.get():
                if row[1].value==hashlib.sha512(password.encode()).hexdigest():
                    status=True
                    if row[7].value=="Admin":
                        self.destroy()
                        Dashboard(row[2].value, row[7].value,row[0].value)
                        excelfile.close()
                    else:
                        self.destroy()
                        MainSession(row[2].value, row[7].value,row[0].value)
                        excelfile.close()
                    break
                else:
                    break
        if status==False:
            ask=mb.askyesno("Thông báo","Mật khẩu hoặc tên đăng nhập không hợp lệ. Nếu quên mật khẩu, bấm Yes.",master=self)
            if ask==True:
                ForgetPassword()
            if ask==False:
                pass
class ForgetPassword(tk.Tk):
    def __init__(self):
        global email_entry
        tk.Tk.__init__(self)
        # window
        self.geometry("400x200+400+400")
        self.title("OTP")
        self.resizable(0, 0)
        canvas=Canvas(master=self,width=400,height=200)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="forgetpassword.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        email_entry = Entry(master=self,relief="flat",font="Montserrat")
        email_entry.place(x=80, y=92,height=24,width=292)
        Check_button = Button(self, command=self.sendOTP)
        checkbutton_img=PhotoImage(master=canvas,file="receiveOTP.png")
        Check_button.configure(image=checkbutton_img)
        Check_button.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                       foreground="#ffffff", background="#ffffff", borderwidth="0")
        Check_button.place(x=135, y=150)
        self.mainloop()
    def sendOTP(self):
        global email
        status=False
        excelfile=load_workbook("UsersData.xlsx")
        for row in excelfile.worksheets[0]:
            if row[5].value==str(email_entry.get()):
                email=str(email_entry.get())
                status=True
                otp=GenerateOTP(email_entry.get())
                self.destroy()
                OTP_window(otp)
                excelfile.close()
        if status==False:
            mb.showerror("Thông báo", "Email bạn nhập không tồn tại trong hệ thống",master=self)
class OTP_window(tk.Tk):
    def __init__(self,otp):
        global OTP_entry
        tk.Tk.__init__(self)
        # window
        self.geometry("400x200+400+400")
        self.title("OTP")
        self.resizable(0, 0)
        canvas=Canvas(master=self,width=400,height=200)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="otp.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        OTP_entry = Entry(master=self,relief="flat",font="Montserrat")
        OTP_entry.place(x=80, y=92,height=24,width=292)
        Check_button = Button(self, command=lambda:self.checkOTP(otp))
        checkbutton_img=PhotoImage(master=canvas,file="OTPCheck.png")
        Check_button.configure(image=checkbutton_img)
        Check_button.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                       foreground="#ffffff", background="#ffffff", borderwidth="0")
        Check_button.place(x=135, y=150)
        self.mainloop()
    def checkOTP(self, otp):
        if otp==OTP_entry.get():
            excelfile=load_workbook("UsersData.xlsx")
            for row in excelfile.worksheets[0]:
                if row[5].value==email:
                    ChangePassword()
                    excelfile.close()
                    self.destroy()
                    break
        else:
            mb.showerror("Thông báo","OTP bạn nhập không khớp")
class ChangePassword(tk.Tk):
    def __init__(self):
        global new_password_entry,new_password_entry2
        tk.Tk.__init__(self)
        self.geometry("400x200")
        self.resizable(0, 0)
        self.title("Đổi mật khẩu")
        canvas=Canvas(master=self,width=400,height=200)
        bg_img=PhotoImage(master=canvas,file="newpassword.png")
        bg=Label(self,image=bg_img)
        bg.place(x=0,y=0)
        button_img=PhotoImage(master=canvas,file="Save_delivery.png")
        save_button=Button(master=self,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0",command=self.CheckValid)
        save_button.configure(image=button_img)
        save_button.place(x=134,y=150)
        new_password_entry=Entry(self,font="Montserrat",relief="flat",show="-")
        new_password_entry.place(x=97,y=62,width=283,height=25)
        new_password_entry2 = Entry(self, font="Montserrat", relief="flat",show="-")
        new_password_entry2.place(x=97, y=99, width=283, height=25)
        self.mainloop()
    def CheckValid(self):
        excelfile=load_workbook("UsersData.xlsx")
        new_password=new_password_entry.get()
        new_password2=new_password_entry2.get()
        if new_password==new_password2:
            for row in excelfile.worksheets[0]:
                if row[5].value==email:
                    row[1].value=hashlib.sha512(new_password.encode()).hexdigest()
                    excelfile.save("UsersData.xlsx")
                    excelfile.close()
                    self.destroy()
                    mb.showinfo("Thông báo","Thay đổi mật khẩu thành công")
                    break
        else:
            mb.showerror("Thông báo","Mật khẩu nhập lại không khớp",master=self)
InitialSetup()
Log_in()