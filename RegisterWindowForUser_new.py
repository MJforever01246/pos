from tkinter import *
import tkinter as tk
from tkinter import messagebox as mb
import hashlib
from OTP import GenerateOTP
from Mail_User_Profile import SendUserProfile
from openpyxl import load_workbook
class RegisterWindowForUser(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        global name_entry, dob, option_gender,option_position, email_entry, phone_entry, username_entry, password_entry, position_combobox,status_hide,unhide_img,hide_img,hide,RadioButtonAdmin,RadioButtonCashier
        # create window
        self.geometry("1150x600+400+400")
        self.title("Đăng kí tài khoản")
        self.resizable(0, 0)
        back = Frame(master=self, bg='white')
        back.pack_propagate(0)
        back.pack(expand=1)
        canvas=Canvas(master=self,width=1150,height=600)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="registerwindow.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        # name
        name_entry = Entry(self, width=55,relief="flat")
        name_entry.place(x=85, y=126)
        # DOB
        dob = Entry(self,width=55,relief="flat")
        dob.place(x=85, y=215)
        # Gender
        option_gender=StringVar()
        img_male=PhotoImage(master=canvas,file="RadioButtonMale.png")
        RadioButtonMale=Checkbutton(master=self,image=img_male,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nam",command=lambda:self.Checkboxgender("Nam"))
        RadioButtonMale.place(x=140, y=299)
        img_female=PhotoImage(master=canvas,file="RadioButtonFemale.png")
        RadioButtonFemale=Checkbutton(master=self,image=img_female,bg="#ffffff",variable=option_gender,cursor="hand2",activebackground="#ffffff",onvalue="Nữ",command=lambda:self.Checkboxgender("Nữ"))
        RadioButtonFemale.place(x=276, y=299)
        # email
        email_entry = Entry(self, width=55,relief="flat")
        email_entry.place(x=703, y=126)
        # phone
        phone_entry = Entry(self, width=55,relief="flat")
        phone_entry.place(x=703, y=215)
        # username
        username_entry = Entry(self, width=55,relief="flat")
        username_entry.place(x=703, y=305)
        # password
        password_entry = Entry(self, width=55, show="-",relief="flat")
        password_entry.place(x=703, y=393)
        hide_img=PhotoImage(master=canvas,file="HidePassword.png")
        unhide_img=PhotoImage(master=canvas,file="UnhidePassword.png")
        hide=Button(self,command=self.hide_unhide_password,image=hide_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        status_hide=True
        hide.place(x=1059,y=389)
        # position
        option_position=tk.StringVar()
        admin_img=PhotoImage(master=canvas,file="RadioButtonAdmin.png")
        RadioButtonAdmin=Checkbutton(master=self,image=admin_img,bg="#ffffff",variable=option_position,cursor="hand2",activebackground="#ffffff",command=lambda:self.Checkboxposition("Admin"),onvalue="Admin")
        RadioButtonAdmin.place(x=140,y=389)
        cashier_img = PhotoImage(master=canvas, file="RadioButtonCashier.png")
        RadioButtonCashier = Checkbutton(master=self, image=cashier_img, bg="#ffffff", variable=option_position,cursor="hand2",activebackground="#ffffff",command=lambda:self.Checkboxposition("Thu ngân"),onvalue="Thu ngân")
        RadioButtonCashier.place(x=276, y=389)
        # button
        complete_img=PhotoImage(master=canvas,file="Save.png")
        Complete = Button(self,command=self.process,image=complete_img,relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
        foreground="#ffffff", background="#ffffff", borderwidth="0")
        Complete.place(x=505, y=550)
        self.mainloop()
    def hide_unhide_password(self):
        global status_hide
        if status_hide==True:
            hide.configure(image=unhide_img)
            password_entry.configure(show="")
            status_hide=False
        else:
            hide.configure(image=hide_img)
            password_entry.configure(show="-")
            status_hide = True
    def Checkboxposition(self,option):
        option_position.set(option)
    def Checkboxgender(self,option):
        option_gender.set(option)
    def process(self):
        global email_info, phone_info, status, otp, window_otp,gender,position
        status = True
        email_info = email_entry.get()
        phone_info = phone_entry.get()
        status_email = True
        status_phone = True
        status_username = True
        status_position = True
        status_password = True
        status_dob = True
        excelfile = load_workbook("UsersData.xlsx")
        user_data = excelfile.worksheets[0]
        for row in user_data.rows:
            if row[0].value == username_entry.get():
                status = False
                status_username = False
            if row[5].value == email_info:
                status = False
                status_email = False
            if row[6].value == phone_info:
                status = False
                status_phone = False
        excelfile.close()
        if status_email == True:
            if email_info.find("@") == -1 or email_info.find(".") == -1:
                status = False
                status_email = False
        if status_phone == True:
            for i in phone_info:
                if i in ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]:
                    pass
                else:
                    status_phone = False
                    status = False
                    break
            if status_phone == True:
                if len(phone_info) == 10 and phone_info[0] == "0":
                    pass
                else:
                    status = False
                    status_phone = False
        if status_username == True:
            if len(username_entry.get()) <= 3:
                status = False
                status_username = False
        if dob.get()=="":
            pass
        else:
            cache=dob.get()
            try:
                cache=cache.split("/")
                if int(cache[0]) > 0 and int(cache[0]) <= 31:
                    pass
                elif int(cache[1]) > 0 and int(cache[1]) <=12:
                    pass
                elif len(cache[2]) == 4:
                    pass
                else:
                    status=False
                    status_dob=False
            except:
                status_dob=False
                status=False
        if option_position.get()=="Admin" or option_position.get()=="Thu ngân":
            pass
        else:
            status_position=False
            status=False
        if password_entry.get() == "":
            status = False
            status_password = False
        if status == False:
            info = ""
            if status_dob == False:
                if info == "":
                    info+="Ngày sinh"
                else:
                    info+=", Ngày sinh"
            if status_position == False:
                if info == "":
                    info += "Chức vụ"
                else:
                    info += ", Chức vụ"
            if status_phone == False:
                if info == "":
                    info += "Di động"
                else:
                    info += ", Di động"
            if status_email == False:
                if info == "":
                    info += "Email"
                else:
                    info += ", Email"
            if status_username == False:
                if info == "":
                    info += "Tên đăng nhập"
                else:
                    info += ", Tên đăng nhập"
            if status_password == False:
                if info == "":
                    info += "Mật khẩu"
                else:
                    info += ", Mật khẩu"
            mb.showerror("Thông báo", info + " không hợp lệ",master=self)
        if status == True:
            try:
                otp = GenerateOTP(email_info)
                window_otp = OTP_window(otp,str(username_entry.get()), str(password_entry.get()),str(name_entry.get()),str(dob.get()),str(option_gender.get()),email_info,phone_info, str(option_position.get()),self)
            except:
                mb.showerror("Thông báo", "Kiểm tra lại email đã nhập hoặc mạng internet", master=self)
class OTP_window(tk.Tk):
    def __init__(self,otp,username,password,name,dob,gender,email,phone,position,window):
        global OTP_entry, window_otp
        tk.Tk.__init__(self)
        # window
        self.geometry("400x200+400+400")
        self.title("OTP")
        self.resizable(0, 0)
        canvas=Canvas(master=self,width=400,height=200)
        bg=Label(self)
        bg_img=PhotoImage(master=canvas,file="otp.png")
        bg.configure(image=bg_img)
        bg.place(relx=0,rely=0)
        OTP_entry = Entry(master=self,relief="flat")
        OTP_entry.place(x=80, y=92,height=24,width=292)
        Check_button = Button(self, command=lambda:self.checkOTP(otp,username,password,name,dob,gender,email,phone,position,window))
        checkbutton_img=PhotoImage(master=canvas,file="OTPCheck.png")
        Check_button.configure(image=checkbutton_img)
        Check_button.configure(relief="flat", overrelief="flat", activebackground="#ffffff", cursor="hand2",
                       foreground="#ffffff", background="#ffffff", borderwidth="0")
        Check_button.place(x=135, y=150)
        self.mainloop()
    def checkOTP(self,otp,username,password,name,dob,gender,email,phone,position,window):
        if OTP_entry.get() == otp:
            try:
                SendUserProfile(name,dob,gender,email,phone,username,position)
            except:
                pass
            self.SaveUserInfo(username,password,name,dob,gender,email,phone,position,window)
            self.destroy()
        else:
            mb.askyesno("Thông báo", "Mã OTP của bạn không trùng khớp. Bạn có muốn thử lại hay không?",master=self)
            if "yes":
                pass
            if "no":
                window.destroy()
                self.destroy()
                mb.showerror("Thông báo","Đăng kí tài khoản không thành công")
    def SaveUserInfo(self,username,password,name,dob,gender,email,phone,position,window):
        excelfile = load_workbook("UsersData.xlsx")
        data=excelfile["UsersData"]
        savedata = [username,str(hashlib.sha512(password.encode())),name,dob,gender,email,phone,position]
        data.insert_rows(idx=2, amount=1)
        list_order = ["A2", "B2", "C2", "D2", "E2", "F2", "G2", "H2"]
        for i in range(8):
            data[list_order[i]] = savedata[i]
        excelfile.save("UsersData.xlsx")
        mb.showinfo("Thông báo","Thêm nhân viên thành công",master=window)
        window.destroy()
        excelfile.close()
def OpenRegisterWindow():
    window_register=RegisterWindowForUser()
# OpenRegisterWindow()
# otp=OTP_window("2561","ádf","dsfad","ádfdf","sadfasdf","ádfasdf","ádfsdaf","sdfasdf","ádfasdf","Ádfd")